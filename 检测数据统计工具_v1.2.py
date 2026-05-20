#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
简易批号计数统计工具
-----------------------------------
功能：
1. 使用 Tkinter 提供简洁的桌面界面
2. 支持一次性导入多个 .xlsx 文件
3. 自动遍历每个工作表，统计每天每个工作表的批号去重数量
4. 汇总所有工作表的每日批号总数
5. 使用 pandas 读取 Excel，openpyxl 写入结果
6. 支持中文显示，具备基本错误处理
"""

# Version: 1.2
# 修改说明：
# - 为导出 Excel 添加 SUM 公式计算合计行
# - 为所有单元格添加细边框
# - 合计行字体加粗
# - 修复了 try/except 缩进导致的 SyntaxError
# - 在代码中加入版本号和修改说明注释
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side
from typing import Dict



# --------------------------- 配置区 ---------------------------
# 需要的关键列（按文件实际列顺序）
DATE_COL = "日期"
BATCH_COL = "批号"
# 读取 Excel 时跳过前两行（表头信息）
SKIP_ROWS = 2
# --------------------------- 工具函数 ---------------------------
VERSION = "1.2"  # 代码版本号，记录本次修改
AUTHOR = "HotLL"
RELEASE_DATE = "2026-05-20"

def _ensure_date_str(val) -> str:
    """统一转为 YYYY-MM-DD 字符串（处理 Excel 序列号、datetime、字符串）。"""
    import datetime as _dt

    if pd.isna(val):
        return ""
    # --- 已是 datetime / Timestamp ---
    if isinstance(val, (pd.Timestamp, _dt.datetime, _dt.date)):
        return val.strftime("%Y-%m-%d")
    # --- 转字符串 ---
    s = str(val).strip()
    if not s or s.lower() in ("nat", "nan", "none", "null", "na"):
        return ""
    #       # 尝试作为 Excel 序列号（数值或数值型字符串）
    try:
        num = float(s)
        # 合理序列号范围（1~200000）
        if 0 <= num <= 200000:
            dt = pd.Timestamp("1899-12-30") + pd.Timedelta(days=num)
            if 1900 <= dt.year <= 2100:
                return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass

    # 解析 YYYYMMDD 整数格式（如 20250804 表示 2025-08-04）
    try:
        num = float(s)
        if 20250101 <= num <= 21000101:
            s_int = str(int(num))
            if len(s_int) == 8:
                dt = pd.to_datetime(s_int, format='%Y%m%d', errors='coerce')
                if pd.notna(dt):
                    return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass

    # 尝试作为字符串日期解析（ISO、中文、/ 分隔等）
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    # 兜底
    return ""



def read_sheet(file_path: str, sheet_name: str) -> pd.DataFrame:
    """
    读取单个工作表，返回仅包含日期（YYYY-MM-DD）和批号的 DataFrame。
    兼容 Excel 序列号、datetime 对象、字符串等多种日期表示。
    """
    try:
        # 不强制指定 dtype，让 pandas 自动识别类型（数值/字符串/datetime）
        # _ensure_date_str 会逐一处理各种类型
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            skiprows=SKIP_ROWS,
            engine="openpyxl",
        )
    except Exception as e:
        raise RuntimeError(f"读取 {os.path.basename(file_path)} - {sheet_name} 失败: {e}")

    # 取前两列作为日期、批号，重命名
    if df.shape[1] < 2:
        # 工作表列数不足，返回空表
        return pd.DataFrame(columns=[DATE_COL, BATCH_COL])
    df = df.iloc[:, :2].copy()
    df.columns = [DATE_COL, BATCH_COL]
    # 逐行转换日期
    df[DATE_COL] = df[DATE_COL].apply(_ensure_date_str)
    # 将批号统一转为字符串（便于后续过滤和统计）
    df[BATCH_COL] = df[BATCH_COL].astype(str).str.strip()
    # Excel 中同一天的后续行日期通常为空单元格，需要前向填充
    df[DATE_COL] = df[DATE_COL].replace("", pd.NA).ffill()
    # 前向填充空日期（Excel 中同一天的后续行日期通常为空）
    df[DATE_COL] = df[DATE_COL].replace("", pd.NA).ffill()
    # 只保留标准日期字符串格式 YYYY-MM-DD 的行
    df = df[df[DATE_COL].str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)]
    # 去除批号为示例的行
    df = df[~df[BATCH_COL].str.startswith("例")]
    return df[[DATE_COL, BATCH_COL]]

def count_daily_batches(df: pd.DataFrame):
    """返回 Series：索引为日期，值为该日期批号去重计数。"""
    daily_counts = df.groupby(DATE_COL)[BATCH_COL].nunique()
    daily_counts.name = "count"
    return daily_counts

def merge_daily_counts(count_dict: Dict[str, pd.Series]) -> pd.DataFrame:
    """将各工作表的每日计数合并为宽表，列名为工作表名称。"""
    frames = []
    for sheet, series in count_dict.items():
        s = series.copy()
        s.name = sheet
        frames.append(s)
    if not frames:
        return pd.DataFrame()
    df_merged = pd.concat(frames, axis=1)
    df_merged = df_merged.fillna(0).astype(int)
    df_merged.index.name = DATE_COL
    return df_merged

# --------------------------- 主界面类 ---------------------------
class BatchCountApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("检测数据统计工具")
        self.geometry("900x600")
        self.resizable(True, True)
        self.file_paths: list[str] = []  # 已导入文件路径列表
        self.results: pd.DataFrame = pd.DataFrame()   # 最终统计 DataFrame（宽表）
        self.total_series: pd.Series = pd.Series(dtype=int)  # 每日汇总（行合计）
        self.column_totals: pd.Series = pd.Series(dtype=int)  # 每列合计（供底行显示）
        self.file_stats: dict[str, dict[str, pd.Series]] = {}  # 每个文件各自的统计字典
        self.all_sheets: list[str] = []  # 所有文件中找到的工作表名称（去重）
        self.create_widgets()
        self._create_menu()
        # 在主窗口底部显示版本号
        self.version_label = tk.Label(self, text=f"Version: {VERSION} | Author: {AUTHOR} | Release: {RELEASE_DATE}", anchor="e")
        self.version_label.pack(side="bottom", fill="x")

    # ---------- UI ----------
    def _setup_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("vista")
        except tk.TclError:
            pass
        style.configure("TButton", background="#87CEFA", foreground="#000000", padding=(14, 7), font=("Microsoft YaHei UI", 10))
        style.configure("Sidebar.TButton", background="#87CEFA", foreground="#000000", padding=(16, 9), font=("Microsoft YaHei UI", 10),
                        anchor="center")
        style.configure("TLabelFrame.Label", background="#87CEFA", foreground="#000000", font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("Treeview", background="#E0F7FF", foreground="#000000", fieldbackground="#E0F7FF", font=("Microsoft YaHei UI", 10), rowheight=26)
        style.configure("Treeview.Heading", background="#87CEFA", foreground="#000000", font=("Microsoft YaHei UI", 10, "bold"))
        style.map("Sidebar.TButton",
                  background=[("active", "#6CA6CD")])

    def create_widgets(self):
        self._setup_style()

        # ========== 左侧功能栏 ==========
        sidebar = ttk.Frame(self, width=150, relief="flat")
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # 操作标题
        ttk.Label(sidebar, text=" 操作", font=("Microsoft YaHei UI", 11, "bold")).pack(
            anchor="w", pady=(18, 10), padx=6)

        btn_cfg = {"side": "top", "fill": "x", "padx": 10, "pady": 3}
        ttk.Button(sidebar, text="导入文件", style="Sidebar.TButton",
                   command=self.import_files).pack(**btn_cfg)
        ttk.Button(sidebar, text="批量导出", style="Sidebar.TButton",
                   command=self.batch_export).pack(**btn_cfg)
        ttk.Button(sidebar, text="打开输出目录", style="Sidebar.TButton",
                   command=self.open_export_folder).pack(**btn_cfg)
        ttk.Button(sidebar, text="清空列表", style="Sidebar.TButton",
                   command=self.clear_all).pack(**btn_cfg)

        # 分隔线
        ttk.Separator(self, orient="vertical").pack(side="left", fill="y", padx=0)

        # ========== 右侧内容区 ==========
        right_frame = ttk.Frame(self)
        right_frame.pack(side="right", fill="both", expand=True)

        # 已导入文件列表
        file_frame = ttk.LabelFrame(right_frame, text="已导入文件")
        file_frame.pack(fill="x", padx=12, pady=(12, 6))
        self.file_listbox = tk.Listbox(file_frame, height=5,
                                       font=("Microsoft YaHei UI", 10), relief="flat",
                                       borderwidth=1, highlightthickness=0)
        self.file_listbox.pack(fill="both", expand=True, side="left", padx=6, pady=6)
        sb_file = ttk.Scrollbar(file_frame, command=self.file_listbox.yview)
        sb_file.pack(side="right", fill="y", pady=6, padx=(0, 4))
        self.file_listbox.configure(yscrollcommand=sb_file.set)
        self.file_listbox.bind('<<ListboxSelect>>', self.on_file_select)

        # 统计结果预览区
        result_frame = ttk.LabelFrame(right_frame, text="统计结果预览（表格形式）")
        result_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))
        self.tree = ttk.Treeview(result_frame, show="headings")
        self.tree.pack(fill="both", expand=True, side="left", padx=(6, 0), pady=6)
        sb_tree = ttk.Scrollbar(result_frame, command=self.tree.yview)
        sb_tree.pack(side="right", fill="y", pady=6, padx=(0, 4))
        self.tree.configure(yscrollcommand=sb_tree.set)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(right_frame, textvariable=self.status_var,
                               relief="sunken", anchor="w",
                               font=("Microsoft YaHei UI", 9))
        status_bar.pack(fill="x", side="bottom", padx=12, pady=(0, 8))

    # ---------- 事件处理 ----------
    def _create_menu(self):
        menubar = tk.Menu(self)
        menubar.add_command(label="关于", command=self.show_about)
        self.config(menu=menubar)

    def show_about(self):
        info = f"版本: {VERSION}\n作者: {AUTHOR}\n发布日期: {RELEASE_DATE}"
        messagebox.showinfo("关于", info)
    def set_status(self, msg: str):
        self.status_var.set(msg)
        self.update_idletasks()

    def import_files(self):
        paths = filedialog.askopenfilenames(
            title="请选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not paths:
            return
        self.file_paths = list(paths)
        self.file_listbox.delete(0, tk.END)
        for p in self.file_paths:
            self.file_listbox.insert(tk.END, os.path.basename(p))
        self.set_status(f"已导入 {len(self.file_paths)} 个文件，开始统计…")
        self.process_files()

    def clear_all(self):
        self.file_paths = []
        self.file_listbox.delete(0, tk.END)
        self.tree.delete(*self.tree.get_children())
        self.results = pd.DataFrame()
        self.total_series = pd.Series(dtype=int)
        self.column_totals = pd.Series(dtype=int)
        self.file_stats.clear()
        self.all_sheets.clear()
        self._merged_results = pd.DataFrame()
        self._merged_total = pd.Series(dtype=int)
        self._merged_column_totals = pd.Series(dtype=int)
        self.set_status("已清空所有数据")

    def on_file_select(self, event):
        selection = self.file_listbox.curselection()
        if not selection:
            # 取消选择时恢复显示合并结果
            self.results = self._merged_results
            self.total_series = self._merged_total
            self.column_totals = self._merged_column_totals
            if not self.results.empty:
                self.display_results()
                self.set_status("显示: 全部文件统计结果")
            return
        idx = selection[0]
        file_path = self.file_paths[idx]
        file_sheet = self.file_stats.get(file_path, {})
        if not file_sheet:
            self.set_status(f"文件 {os.path.basename(file_path)} 无统计结果")
            return
        # 计算单个文件的统计结果（与 batch_export 逻辑一致）
        file_results = merge_daily_counts(file_sheet)
        file_sheet_names = list(file_sheet.keys())
        file_results = file_results.reindex(columns=file_sheet_names, fill_value=0)
        file_results.sort_index(inplace=True)
        self.results = file_results
        self.total_series = file_results.sum(axis=1).astype(int)
        self.column_totals = file_results.sum().astype(int)
        self.display_results()
        self.set_status(f"显示: {os.path.basename(file_path)}")

    def process_files(self):
        if not self.file_paths:
            messagebox.showwarning("提示", "请先导入 Excel 文件")
            return
        # 统计每个工作表、每天的批号去重计数
        sheet_counts = {}  # 全局合并统计 {sheet_name: Series(date->count)}
        self.file_stats.clear()
        for file_path in self.file_paths:
            try:
                xl = pd.ExcelFile(file_path, engine="openpyxl")
            except Exception as e:
                messagebox.showerror("读取错误", f"文件 {os.path.basename(file_path)} 读取失败: {e}")
                continue
            file_sheet = {}  # 当前文件的统计
            for sheet_name in xl.sheet_names:
                if sheet_name == "目录":
                    continue
                try:
                    df_sheet = read_sheet(file_path, sheet_name)
                except Exception as e:
                    messagebox.showerror("读取错误", f"{os.path.basename(file_path)} - {sheet_name}: {e}")
                    continue
                if df_sheet.empty:
                    continue
                daily_counts = count_daily_batches(df_sheet)
                # 记录到当前文件
                file_sheet[sheet_name] = daily_counts.astype(int)
                # 合并到全局
                if sheet_name in sheet_counts:
                    sheet_counts[sheet_name] = sheet_counts[sheet_name].add(daily_counts, fill_value=0).astype(int)
                else:
                    sheet_counts[sheet_name] = daily_counts.astype(int)
            self.file_stats[file_path] = file_sheet
        # 收集所有出现的工作表名称（按首次出现的顺序）
        self.all_sheets = list(sheet_counts.keys())
        # 合并为宽表
        self.results = merge_daily_counts(sheet_counts)
        # 确保所有出现的工作表列都存在，缺失的填充为0
        self.results = self.results.reindex(columns=self.all_sheets, fill_value=0)
        # 按日期升序排序
        self.results.sort_index(inplace=True)
        if self.results.empty:
            self.set_status("未检测到有效数据")
            return
        # 计算每日汇总（所有工作表合计）
        self.total_series = self.results.sum(axis=1).astype(int)
        # 计算每列合计（用于底行显示）
        self.column_totals = self.results.sum().astype(int)
        # 保存合并结果的副本（供单个文件选择时恢复）
        self._merged_results = self.results.copy()
        self._merged_total = self.total_series.copy()
        self._merged_column_totals = self.column_totals.copy()
        # 在 UI 中展示
        self.display_results()
        self.set_status("统计完成")

    def display_results(self):
        # 清空已有树
        self.tree.delete(*self.tree.get_children())
        # 设置列：日期 + 工作表列 + 合计列
        columns = ["日期"] + list(self.results.columns) + ["合计"]
        self.tree["columns"] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor="center")
        # 填充数据行
        for date, row in self.results.iterrows():
            values = [date] + list(row) + [self.total_series.get(date, 0)]
            self.tree.insert("", tk.END, values=values)
        # 添加底行合计
        if not self.column_totals.empty:
            total_vals = ["合计"] + list(self.column_totals) + [self.column_totals.sum()]
            self.tree.insert("", tk.END, values=total_vals, tags=("total",))
        self.tree.tag_configure("total", background="#f0f0f0", font=("", 10, "bold"))

    def _write_formatted_excel(self, save_path, results, total_series, column_totals):
        """将统计结果写入 Excel 并设置格式（通用方法）。"""
        # 若目标文件已存在且被占用，先尝试删除后再写入
        if os.path.exists(save_path):
            try:
                os.remove(save_path)
            except PermissionError:
                messagebox.showwarning("文件占用", f"文件 {os.path.basename(save_path)} 正在使用，请先关闭或选择其他保存位置。")
                return False
        try:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                # ---------- 每日批号统计 ----------
                detail_df = results.reset_index()
                # 将 0 值替换为空（不显示），保留日期列
                for col in detail_df.columns:
                    if col != "日期":
                        detail_df[col] = detail_df[col].replace(0, None)
                detail_df.to_excel(writer, sheet_name="每日批号统计", index=False)
                # 添加合计行，使用 Excel SUM 公式
                ws_detail = writer.sheets["每日批号统计"]
                start_row = 2  # 数据起始行（含标题行后第一行）
                end_row = ws_detail.max_row
                total_row = end_row + 1
                ws_detail.cell(row=total_row, column=1, value="合计")
                for col_idx, col_name in enumerate(detail_df.columns[1:], start=2):
                    col_letter = ws_detail.cell(row=1, column=col_idx).column_letter
                    formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                    ws_detail.cell(row=total_row, column=col_idx, value=formula)
                # ---------- 每日汇总 ----------
                total_df = total_series.rename("合计").to_frame().reset_index()
                # 将 0 值替换为空
                total_df["合计"] = total_df["合计"].replace(0, None)
                total_df.to_excel(writer, sheet_name="每日汇总", index=False)
                # 添加合计行，使用 Excel SUM 公式
                ws_total = writer.sheets["每日汇总"]
                start_row = 2
                end_row = ws_total.max_row
                total_row = end_row + 1
                ws_total.cell(row=total_row, column=1, value="合计")
                col_letter = ws_total.cell(row=1, column=2).column_letter
                formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                ws_total.cell(row=total_row, column=2, value=formula)
                # ---------- 统一设置格式 ----------
                font_style = Font(name="仿宋", size=14)
                align_center = Alignment(horizontal="center", vertical="center")
                for ws in writer.sheets.values():
                    thin = Side(style="thin")
                    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    # 记录总行列位置（假设最后一行/列为合计）
                    total_row = ws.max_row
                    total_col = ws.max_column
                    for col_cells in ws.columns:
                        max_len = 0
                        col_letter = col_cells[0].column_letter
                        for cell in col_cells:
                            # 基础样式
                            cell.font = font_style
                            cell.alignment = align_center
                            cell.border = thin_border
                            # 合计行或合计列加粗
                            if cell.row == total_row or cell.column == total_col:
                                cell.font = Font(name="仿宋", size=14, bold=True)
                            val = str(cell.value) if cell.value is not None else ""
                            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                            max_len = max(max_len, char_len + 2)
                        ws.column_dimensions[col_letter].width = max(min(max_len + 2, 30), 8)
                    for row_dim in ws.row_dimensions.values():
                        row_dim.height = 35
            return True
        except Exception as e:
            messagebox.showerror("导出错误", f"导出失败: {e}")
            return False

    def _get_output_dir(self) -> str:
        """获取导入文件目录下的 统计后结果 文件夹路径（不存在则创建）。"""
        if not self.file_paths:
            return ""
        base_dir = os.path.dirname(self.file_paths[0])
        out_dir = os.path.join(base_dir, "统计后结果")
        os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def open_export_folder(self):
        """在文件管理器中打开 统计后结果 文件夹。"""
        out_dir = self._get_output_dir()
        if not out_dir or not os.path.exists(out_dir):
            messagebox.showinfo("提示", "尚未导出过结果，请先执行批量导出。")
            return
        os.startfile(out_dir)
        self.set_status(f"已打开文件夹: {out_dir}")

    def batch_export(self):
        """为每个已导入文件生成独立的统计结果文件（文件名 + _统计结果.xlsx）。"""
        if not self.file_paths or not self.file_stats:
            messagebox.showwarning("提示", "请先导入并统计文件，再执行批量导出。")
            return
        # 确认对话框
        if not messagebox.askyesno("批量导出", f"将为 {len(self.file_paths)} 个文件分别生成统计结果，是否继续？"):
            return
        success_count = 0
        self.set_status(f"正在批量导出（共 {len(self.file_paths)} 个文件）…")
        self.update_idletasks()
        for file_path in self.file_paths:
            file_sheet = self.file_stats.get(file_path, {})
            if not file_sheet:
                continue
            # 对该文件单独统计
            file_results = merge_daily_counts(file_sheet)
            # 只保留当前文件中实际已有的工作表列
            file_sheet_names = list(file_sheet.keys())
            file_results = file_results.reindex(columns=file_sheet_names, fill_value=0)
            file_results.sort_index(inplace=True)
            file_total = file_results.sum(axis=1).astype(int)
            file_col_total = file_results.sum().astype(int)
            # 生成输出文件名：原文件名_统计结果.xlsx，放入 统计后结果 文件夹
            out_dir = self._get_output_dir()
            if not out_dir:
                continue
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            out_path = os.path.join(out_dir, f"{base_name}_统计结果.xlsx")
            ok = self._write_formatted_excel(out_path, file_results, file_total, file_col_total)
            if ok:
                success_count += 1
        messagebox.showinfo("批量导出完成", f"成功导出 {success_count} / {len(self.file_paths)} 个文件")
        self.set_status(f"批量导出完成，成功 {success_count} 个文件")


# --------------------------- 程序入口 ---------------------------
if __name__ == "__main__":
    # 兼容 Windows 中文控制台
    import sys
    # Windows 环境下默认使用 UTF-8 编码，无需额外设置
    app = BatchCountApp()
    app.mainloop()
# Version: 1.2
# 本次修改说明：
# - 为导出 Excel 添加 SUM 公式计算合计行
# - 为所有单元格添加细边框
# - 合计行字体加粗
# - 修复了 try/except 缩进导致的 SyntaxError
# - 在代码中加入版本号和修改说明注释