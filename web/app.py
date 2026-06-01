#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
检测数据统计工具 Web 版
基于 v2.2.10 核心逻辑重构为 Flask Web 应用
"""

import os
import sys
import json
import io
import zipfile
import logging
import tempfile
import shutil
import datetime as _dt
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from flask import Flask, render_template, request, jsonify, send_file

# --------------------------- 配置 ---------------------------
DATE_COL: str = "日期"
BATCH_COL: str = "批号"
SKIP_ROWS: int = 2
VERSION: str = "3.0.0"
AUTHOR: str = "HotLL"

BATCH_EXAMPLE_PREFIXES: Tuple[str, ...] = ("例", "示例", "sample")

# --------------------------- 日志 ---------------------------
LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "web_count.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)

# --------------------------- Flask ---------------------------
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 500 * 1024 * 1024  # 500MB

# --------------------------- 核心逻辑（复用 v2.2.10） ---------------------------
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side


def _ensure_date_str(val) -> str:
    """统一转为 YYYY-MM-DD 字符串（支持 YYYYMMDD 格式）"""
    if val is None:
        return ""
    if isinstance(val, str) and val.strip() == "":
        return ""
    try:
        s = str(val).strip()
        if "." in s:
            s = s.split(".")[0]
        if len(s) == 8 and s.isdigit():
            year, month, day = int(s[:4]), int(s[4:6]), int(s[6:8])
            if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                return f"{year:04d}-{month:02d}-{day:02d}"
    except:
        pass
    if isinstance(val, float) and pd.isna(val):
        return ""
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if not pd.isna(dt):
            if dt.year == 1970:
                # 1970 年结果通常是数值被 pd.to_datetime 误解析（如 21.5→1970-01-21）
                # 仅当 val 本身是合法年份整数时才保留
                try:
                    num = float(val)
                    if not (1900 <= num <= 2100):
                        return ""
                except (ValueError, TypeError):
                    return ""
            if 1900 <= dt.year <= 2100:
                return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    return ""


def _detect_data_start_row(df: pd.DataFrame, max_scan: int = 50) -> int:
    """扫描前 max_scan 行 × 前 30 列，返回第一个有效日期所在的行号
    未找到则返回 SKIP_ROWS（保持向后兼容）"""
    import re
    ncols = min(df.shape[1], 30)
    for row_idx in range(min(max_scan, df.shape[0])):
        for col_idx in range(ncols):
            try:
                val = df.iloc[row_idx, col_idx]
            except Exception:
                continue
            date_str = _ensure_date_str(val)
            if date_str and re.match(r"^\d{4}-\d{2}-\d{2}$", date_str):
                logger.debug("自动检测到数据起始行: 第 %d 行 (第 %d 列)", row_idx + 1, col_idx + 1)
                return row_idx
    logger.debug("未检测到有效日期，使用默认跳过 %d 行", SKIP_ROWS)
    return SKIP_ROWS


def _detect_columns(header_df: pd.DataFrame, max_header_rows: int = 3) -> Tuple[int, int]:
    """扫描表头区域，自动识别日期列和批号列的位置
    返回 (date_col, batch_col)，未识别到则回退默认 (0, 1)"""
    date_keywords = [
        "日期", "date", "时间", "time",
        "检测日期", "采样日期", "报告日期", "分析日期", "送检日期", "检验日期",
    ]
    batch_keywords = [
        "批号", "batch", "lot",
        "编号", "序号",
        "样品批号", "样品编号", "样品号", "样品名称",
        "实验编号", "样本号",
        "code", "id",
    ]

    date_col = 0
    batch_col = 1
    date_found = False
    batch_found = False

    if header_df.empty or header_df.shape[1] < 2:
        return date_col, batch_col

    # 从表头区域最后几行向上扫描（最靠近数据的表头行优先）
    start = max(0, header_df.shape[0] - max_header_rows)
    for row_idx in range(header_df.shape[0] - 1, start - 1, -1):
        for col_idx in range(min(header_df.shape[1], 30)):
            try:
                val = str(header_df.iloc[row_idx, col_idx]).strip().lower()
            except Exception:
                continue
            if not val:
                continue
            if not date_found and any(kw in val for kw in date_keywords):
                date_col = col_idx
                date_found = True
            if not batch_found and any(kw in val for kw in batch_keywords):
                batch_col = col_idx
                batch_found = True
        if date_found and batch_found:
            break

    if date_found or batch_found:
        logger.debug(
            "自动识别列: 日期=第%d列, 批号=第%d列 (匹配关键词)",
            date_col + 1, batch_col + 1
        )
    return date_col, batch_col


def read_sheet_data(file_path: str, sheet_name: str) -> Optional[pd.DataFrame]:
    """读取工作表数据，返回 (日期, 批号) DataFrame（对齐 v2.2.10）"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, engine="openpyxl")
        if df.shape[1] < 2:
            return None

        # 自动检测数据起始行：扫描找到第一个有效日期所在行
        auto_skip = _detect_data_start_row(df)

        # 提取表头区域，自动识别日期列和批号列位置
        header_area = df.iloc[:auto_skip] if auto_skip > 0 else df.iloc[:0]
        date_col, batch_col = _detect_columns(header_area)

        # 切除表头，只保留数据行
        if auto_skip > 0:
            df = df.iloc[auto_skip:]
            logger.info(" %s: 自动跳过前 %d 行表头", sheet_name, auto_skip)

        # 选取识别到的日期列和批号列
        df = df.iloc[:, [date_col, batch_col]].copy()
        df.columns = [DATE_COL, BATCH_COL]
        df[DATE_COL] = df[DATE_COL].apply(_ensure_date_str)
        df[BATCH_COL] = df[BATCH_COL].astype(str).str.strip()

        # 前向填充日期（合并单元格场景）
        df[DATE_COL] = df[DATE_COL].replace("", pd.NA).ffill()

        # 过滤无效日期行
        df = df[df[DATE_COL].str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)]

        # 过滤示例/模板行
        mask = pd.Series(True, index=df.index)
        for prefix in BATCH_EXAMPLE_PREFIXES:
            mask = mask & ~df[BATCH_COL].str.startswith(prefix)
        df = df[mask]

        return df if not df.empty else None
    except Exception as e:
        logger.error("读取 %s 失败: %s", sheet_name, e)
        return None


def count_daily_batches(df: pd.DataFrame) -> pd.Series:
    """返回 Series：索引为日期，值为该日期批号去重计数"""
    if df.empty:
        return pd.Series(dtype=int)
    daily_counts = df.groupby(DATE_COL)[BATCH_COL].nunique()
    daily_counts.name = "count"
    return daily_counts


def merge_daily_counts(count_dict: Dict[str, pd.Series]) -> pd.DataFrame:
    """将各工作表的每日计数合并为宽表"""
    if not count_dict:
        return pd.DataFrame()
    all_dates = set()
    for series in count_dict.values():
        all_dates.update(series.index)
    if not all_dates:
        return pd.DataFrame()
    result = pd.DataFrame(index=sorted(all_dates))
    for sheet, series in count_dict.items():
        result[sheet] = series
    result = result.fillna(0).astype(int)
    result.index.name = DATE_COL
    return result


def _df_to_table_json(results: pd.DataFrame, total_series: pd.Series,
                      column_totals: pd.Series, mode: str = "") -> dict:
    """将 DataFrame 转换为前端表格 JSON"""
    if results.empty:
        return {"columns": [], "rows": [], "totals": [], "mode": mode}
    columns = ["日期"] + list(results.columns) + ["合计"]
    rows = []
    for date, row in zip(results.index, results.itertuples()):
        vals = [date]
        for v in row[1:]:
            vals.append("" if pd.isna(v) else int(v))
        tv = total_series.get(date)
        vals.append("" if tv is None or pd.isna(tv) else int(tv))
        rows.append(vals)
    totals = (
        ["合计"]
        + [int(column_totals.get(c, 0)) for c in results.columns]
        + [int(column_totals.sum())]
    )
    return {"columns": columns, "rows": rows, "totals": totals, "mode": mode}


# --------------------------- 全局状态（会话级） ---------------------------
_upload_dir: Optional[str] = None
_file_paths: List[str] = []
_file_stats: Dict[str, Dict[str, pd.Series]] = {}
_merged_results: Optional[pd.DataFrame] = None
_merged_total: Optional[pd.Series] = None
_merged_column_totals: Optional[pd.Series] = None


def _get_upload_dir() -> str:
    global _upload_dir
    if _upload_dir is None or not os.path.exists(_upload_dir):
        _upload_dir = tempfile.mkdtemp(prefix="batch_tool_")
    return _upload_dir


def _clear_session():
    global _file_paths, _file_stats, _merged_results, _merged_total, _merged_column_totals, _upload_dir
    _file_paths = []
    _file_stats = {}
    _merged_results = None
    _merged_total = None
    _merged_column_totals = None
    if _upload_dir and os.path.exists(_upload_dir):
        shutil.rmtree(_upload_dir, ignore_errors=True)
    _upload_dir = None


def _compute_single_file(file_path: str):
    file_sheet = _file_stats.get(file_path, {})
    if not file_sheet:
        return None
    results = merge_daily_counts(file_sheet)
    sheet_names = list(file_sheet.keys())
    results = results.reindex(columns=sheet_names, fill_value=0)
    results.sort_index(inplace=True)
    total = results.sum(axis=1).astype(int)
    col_total = results.sum().astype(int)
    return results, total, col_total


def _rebuild_merged():
    global _merged_results, _merged_total, _merged_column_totals
    sheet_counts: Dict[str, pd.Series] = {}
    for series_dict in _file_stats.values():
        for sheet_name, daily_counts in series_dict.items():
            sheet_counts[sheet_name] = sheet_counts.get(sheet_name, pd.Series(dtype=int)).add(
                daily_counts, fill_value=0
            ).astype(int)
    if not sheet_counts:
        _merged_results, _merged_total, _merged_column_totals = None, None, None
        return
    _merged_results = merge_daily_counts(sheet_counts)
    _merged_results = _merged_results.reindex(columns=list(sheet_counts.keys()), fill_value=0)
    _merged_results.sort_index(inplace=True)
    _merged_total = _merged_results.sum(axis=1).astype(int)
    _merged_column_totals = _merged_results.sum().astype(int)


# --------------------------- API 路由 ---------------------------

@app.route("/")
def index():
    return render_template("index.html", version=VERSION, author=AUTHOR)


@app.route("/api/import", methods=["POST"])
def api_import():
    """上传 Excel 文件并统计"""
    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "未选择任何文件", "files": [], "results": None, "status": "就绪"})

    upload_dir = _get_upload_dir()
    new_paths = []

    for f in files:
        if f.filename == "":
            continue
        # 保存到临时目录
        safe_name = f.filename.replace("\\", "_").replace("/", "_")
        save_path = os.path.join(upload_dir, safe_name)
        # 去重：已存在则跳过
        if save_path in _file_paths:
            continue
        # 重名处理
        counter = 1
        base, ext = os.path.splitext(save_path)
        while os.path.exists(save_path):
            save_path = f"{base}_{counter}{ext}"
            counter += 1
        f.save(save_path)
        new_paths.append(save_path)

    if not new_paths:
        file_names = [os.path.basename(p) for p in _file_paths]
        return jsonify({
            "files": file_names,
            "results": _df_to_table_json(_merged_results, _merged_total, _merged_column_totals, "全部文件")
            if _merged_results is not None else None,
            "status": f"所选文件均已导入，当前共 {len(_file_paths)} 个文件",
        })

    _file_paths.extend(new_paths)
    logger.info("导入 %d 个文件，共 %d 个", len(new_paths), len(_file_paths))

    # 处理每个新文件
    for file_path in new_paths:
        try:
            xl = pd.ExcelFile(file_path, engine="openpyxl")
        except Exception as e:
            logger.error("文件 %s 读取失败: %s", os.path.basename(file_path), e)
            continue

        file_sheet: Dict[str, pd.Series] = {}
        for sheet_name in xl.sheet_names:
            if sheet_name == "目录":
                continue
            df_sheet = read_sheet_data(file_path, sheet_name)
            if df_sheet is None or df_sheet.empty:
                logger.info("跳过 %s（无有效数据）", sheet_name)
                continue
            daily_counts = count_daily_batches(df_sheet).astype(int)
            file_sheet[sheet_name] = daily_counts
            logger.info("已处理 %s: %d 条记录", sheet_name, len(df_sheet))
        logger.info("文件 %s 有效工作表: %s", os.path.basename(file_path), list(file_sheet.keys()))
        if file_sheet:
            _file_stats[file_path] = file_sheet

    _rebuild_merged()

    file_names = [os.path.basename(p) for p in _file_paths]
    result_json = None
    if _merged_results is not None:
        result_json = _df_to_table_json(_merged_results, _merged_total, _merged_column_totals, "全部文件")
    else:
        result_json = None

    return jsonify({
        "files": file_names,
        "results": result_json,
        "status": f"已导入 {len(_file_paths)} 个文件（新增 {len(new_paths)} 个），统计完成",
    })


@app.route("/api/results")
def api_results():
    """获取统计结果"""
    file_index = request.args.get("file_index", -1, type=int)
    if file_index < 0 or file_index >= len(_file_paths):
        if _merged_results is None or _merged_results.empty:
            return jsonify({"columns": [], "rows": [], "totals": [], "mode": "全部文件 — 无数据"})
        return jsonify(_df_to_table_json(_merged_results, _merged_total, _merged_column_totals, "全部文件"))

    file_path = _file_paths[file_index]
    result = _compute_single_file(file_path)
    if result is None:
        return jsonify({"columns": [], "rows": [], "totals": [], "mode": f"{os.path.basename(file_path)} — 无数据"})
    file_results, file_total, file_col_total = result
    return jsonify(_df_to_table_json(file_results, file_total, file_col_total, os.path.basename(file_path)))


@app.route("/api/export", methods=["POST"])
def api_export():
    """批量导出：生成 zip 包含每个文件的统计 Excel"""
    if not _file_paths or not _file_stats:
        return jsonify({"error": "请先导入并统计文件"})

    buf = io.BytesIO()
    success_count = 0
    failed_files: List[str] = []

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in _file_paths:
            result = _compute_single_file(file_path)
            if result is None:
                failed_files.append(os.path.basename(file_path))
                continue
            file_results, file_total, file_col_total = result

            # 内存中生成 Excel
            excel_buf = io.BytesIO()
            _write_formatted_excel_to_buffer(excel_buf, file_results, file_total, file_col_total)
            excel_buf.seek(0)

            base_name = os.path.splitext(os.path.basename(file_path))[0]
            zf.writestr(f"{base_name}_统计结果.xlsx", excel_buf.read())
            success_count += 1

    buf.seek(0)

    if failed_files:
        logger.warning("导出失败: %s", ", ".join(failed_files))

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"检测数据统计结果_{timestamp}.zip",
    )


@app.route("/api/download-merged", methods=["POST"])
def api_download_merged():
    """下载合并统计结果 Excel"""
    if _merged_results is None or _merged_results.empty:
        return jsonify({"error": "暂无统计数据"})

    buf = io.BytesIO()
    _write_formatted_excel_to_buffer(buf, _merged_results, _merged_total, _merged_column_totals)
    buf.seek(0)

    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"检测数据统计_合并结果_{timestamp}.xlsx",
    )


@app.route("/api/clear", methods=["POST"])
def api_clear():
    _clear_session()
    return jsonify({"status": "已清空所有数据"})


@app.route("/api/log")
def api_log():
    """返回审计日志（尾部 500 行，降序）"""
    import collections
    log_path = LOG_DIR / "web_count.log"
    MAX_LINES = 500
    for enc in ("utf-8", "gbk", "cp936"):
        try:
            with open(log_path, "r", encoding=enc) as f:
                lines = collections.deque(f, maxlen=MAX_LINES)
                return "\n".join(reversed(lines))
        except UnicodeDecodeError:
            continue
        except Exception:
            break
    try:
        with open(log_path, "r", encoding="utf-8", errors="replace") as f:
            lines = collections.deque(f, maxlen=MAX_LINES)
            return "\n".join(reversed(lines))
    except Exception as e:
        return f"读取日志文件失败: {e}"


# --------------------------- Excel 生成 ---------------------------

def _write_formatted_excel_to_buffer(
    buf: io.BytesIO, results: pd.DataFrame,
    total_series: pd.Series, column_totals: pd.Series,
):
    """在内存中生成格式化 Excel"""
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ---- 每日批号统计 ----
        detail_df = results.reset_index()
        for col in detail_df.columns:
            if col != "日期":
                detail_df[col] = detail_df[col].replace(0, None)
        detail_df["合计"] = total_series.replace(0, None).values
        detail_df.to_excel(writer, sheet_name="每日批号统计", index=False)
        ws_detail = writer.sheets["每日批号统计"]
        start_row, end_row = 2, ws_detail.max_row
        total_row = end_row + 1
        ws_detail.cell(row=total_row, column=1, value="合计")
        for col_idx in range(2, len(detail_df.columns) + 1):
            col_letter = ws_detail.cell(row=1, column=col_idx).column_letter
            ws_detail.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")

        # ---- 每日汇总 ----
        total_df = total_series.rename("合计").to_frame().reset_index()
        total_df["合计"] = total_df["合计"].replace(0, None)
        total_df.to_excel(writer, sheet_name="每日汇总", index=False)
        ws_total = writer.sheets["每日汇总"]
        start_row, end_row = 2, ws_total.max_row
        total_row = end_row + 1
        ws_total.cell(row=total_row, column=1, value="合计")
        col_letter = ws_total.cell(row=1, column=2).column_letter
        ws_total.cell(row=total_row, column=2, value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")

        # ---- 格式美化 ----
        font_style = Font(name="仿宋", size=14)
        align_center = Alignment(horizontal="center", vertical="center")
        for ws in writer.sheets.values():
            thin = Side(style="thin")
            thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            total_row = ws.max_row
            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    cell.font = font_style
                    cell.alignment = align_center
                    cell.border = thin_border
                    if cell.row == total_row:
                        cell.font = Font(name="仿宋", size=14, bold=True)
                max_len = max(
                    (sum(2 if ord(c) > 127 else 1 for c in (str(cell.value) if cell.value is not None else "")) + 2)
                    for cell in col_cells
                )
                ws.column_dimensions[col_letter].width = max(min(max_len + 2, 30), 8)
            for row_dim in ws.row_dimensions.values():
                row_dim.height = 35


# ========================== 启动 ==========================
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--port", type=int, default=5000)
    args = parser.parse_args()
    logger.info(f"启动 Web 版 v{VERSION} (port={args.port})")
    app.run(debug=args.debug, host="127.0.0.1", port=args.port)
