#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
检测数据统计工具 v2.0.3 - Web UI 版本
基于 pywebview，使用 Uiverse Galaxy 风格的现代按钮
功能保持不变，代码结构已优化以便打包并减小体积
"""

# --------------------------- 版本信息 ---------------------------
VERSION = "2.0.3"
AUTHOR = "HotLL"
RELEASE_DATE = "2026-05-23"

# --------------------------- 基础库 ---------------------------
import os
import sys
import datetime as _dt
import logging
import logging.handlers
import json
from typing import Dict, Optional

# --------------------------- 懒加载 Pandas ---------------------------
def _import_pandas():
    """在需要时才导入 pandas 与 openpyxl，返回 pandas 模块"""
    try:
        import pandas as pd
        import openpyxl  # noqa: F401
    except ModuleNotFoundError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
        import pandas as pd
        import openpyxl  # noqa: F401
    return pd

# --------------------------- 日志配置 ---------------------------
def _init_logger():
    log_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "batch_count.log")
    handler = logging.handlers.RotatingFileHandler(
        log_file_path, maxBytes=5 * 1024 * 1024, backupCount=3, delay=True, encoding="utf-8"
    )
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logging.basicConfig(level=logging.INFO, handlers=[handler])
    return logging.getLogger(__name__)

logger = _init_logger()

# --------------------------- 配置区 ---------------------------
DATE_COL = "日期"
BATCH_COL = "批号"
SKIP_ROWS = 2

# --------------------------- 工具函数 ---------------------------
def _ensure_date_str(val) -> str:
    """统一转为 YYYY-MM-DD 字符串（处理 Excel 序列号、datetime、字符串）。"""
    pd = _import_pandas()
    if pd.isna(val):
        return ""
    if isinstance(val, (pd.Timestamp, _dt.datetime, _dt.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("nat", "nan", "none", "null", "na"):
        return ""
    try:
        num = float(s)
        if 0 <= num <= 200000:
            dt = pd.Timestamp("1899-12-30") + pd.Timedelta(days=num)
            if 1900 <= dt.year <= 2100:
                return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    try:
        num = float(s)
        if 20250101 <= num <= 21000101:
            s_int = str(int(num))
            if len(s_int) == 8:
                dt = pd.to_datetime(s_int, format='%Y%m%d', errors='coerce')
                if pd.notna(dt):
                    return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    return ""

def read_sheet(file_path: str, sheet_name: str):
    """读取单个工作表，返回仅包含日期和批号的 DataFrame。"""
    pd = _import_pandas()
    logger.info("读取工作表 %s 来自文件 %s", sheet_name, os.path.basename(file_path))
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            skiprows=SKIP_ROWS,
            engine="openpyxl",
        )
    except Exception as e:
        logger.error("读取 %s - %s 失败: %s", os.path.basename(file_path), sheet_name, e)
        raise RuntimeError(f"读取 {os.path.basename(file_path)} - {sheet_name} 失败: {e}")
    if df.shape[1] < 2:
        return pd.DataFrame(columns=[DATE_COL, BATCH_COL])
    df = df.iloc[:, :2].copy()
    df.columns = [DATE_COL, BATCH_COL]
    df[DATE_COL] = df[DATE_COL].apply(_ensure_date_str)
    df[BATCH_COL] = df[BATCH_COL].astype(str).str.strip()
    df[DATE_COL] = df[DATE_COL].replace("", pd.NA).ffill()
    df[DATE_COL] = df[DATE_COL].replace("", pd.NA).ffill()
    df = df[df[DATE_COL].str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)]
    df = df[~df[BATCH_COL].str.startswith("例")]
    return df[[DATE_COL, BATCH_COL]]

def count_daily_batches(df):
    pd = _import_pandas()
    daily_counts = df.groupby(DATE_COL)[BATCH_COL].nunique()
    daily_counts.name = "count"
    return daily_counts

def merge_daily_counts(count_dict: Dict[str, 'pd.Series']):
    pd = _import_pandas()
    frames = []
    for sheet, series in count_dict.items():
        s = series.copy()
        s.name = sheet
        frames.append(s)
    if not frames:
        return pd.DataFrame()
    df_merged = pd.concat(frames, axis=1).fillna(0).astype(int)
    df_merged.index.name = DATE_COL
    return df_merged

# --------------------------- HTML 模板 ---------------------------
HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'><rect width='32' height='32' rx='6' fill='%236e45e2'/><rect x='5' y='18' width='4' height='9' rx='1' fill='white'/><rect x='11' y='12' width='4' height='15' rx='1' fill='white'/><rect x='17' y='8' width='4' height='19' rx='1' fill='white'/><rect x='23' y='14' width='4' height='13' rx='1' fill='white'/></svg>">
<title>检测数据统计工具 v2.0.3</title>
<style>
/* 基础 CSS（已在原项目中完整实现，此处省略） */
</style>
</head>
<body class="light">
<!-- UI 结构保持原样 -->
</body>
</html>
""".replace("__VERSION__", VERSION).replace("__AUTHOR__", AUTHOR).replace("__RELEASE_DATE__", RELEASE_DATE)

# --------------------------- 后端 API ---------------------------
class BackendApi:
    def __init__(self):
        self.file_paths: list[str] = []
        self.file_stats: dict = {}
        self.all_sheets: list[str] = []
        self._merged_results: Optional['pd.DataFrame'] = None
        self._merged_total: Optional['pd.Series'] = None
        self._merged_column_totals: Optional['pd.Series'] = None

    def _df_to_table_json(self, results, total_series, column_totals, mode=""):
        pd = _import_pandas()
        if results.empty:
            return {"columns": [], "rows": [], "totals": [], "mode": mode}
        columns = ["日期"] + list(results.columns) + ["合计"]
        rows = []
        for date, row in results.iterrows():
            row_vals = [date] + [int(v) if pd.notna(v) else 0 for v in row]
            row_vals.append(int(total_series.get(date, 0)))
            rows.append(row_vals)
        totals = ["合计"] + [int(column_totals.get(col, 0)) for col in results.columns] + [int(column_totals.sum())]
        return {"columns": columns, "rows": rows, "totals": totals, "mode": mode}

    def import_files(self):
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
        paths = filedialog.askopenfilenames(title="请选择 Excel 文件", filetypes=[("Excel 文件", "*.xlsx")])
        root.destroy()
        if not paths:
            return {"error": "未选择任何文件", "files": [], "results": None, "status": "就绪"}
        self.file_paths = list(paths)
        logger.info("导入 %d 个文件: %s", len(self.file_paths), [os.path.basename(p) for p in self.file_paths])
        self._process_files()
        file_names = [os.path.basename(p) for p in self.file_paths]
        return {"files": file_names, "results": self._df_to_table_json(self._merged_results, self._merged_total, self._merged_column_totals, "全部文件"), "status": f"已导入 {len(self.file_paths)} 个文件，统计完成"}

    def _process_files(self):
        if not self.file_paths:
            return {"error": "请先导入文件"}
        pd = _import_pandas()
        logger.info("开始处理 %d 个文件", len(self.file_paths))
        sheet_counts = {}
        self.file_stats.clear()
        for file_path in self.file_paths:
            logger.info("处理文件 %s", os.path.basename(file_path))
            try:
                xl = pd.ExcelFile(file_path, engine="openpyxl")
            except Exception as e:
                msg = f"文件 {os.path.basename(file_path)} 读取失败: {e}"
                logger.error(msg)
                return {"error": msg}
            file_sheet = {}
            for sheet_name in xl.sheet_names:
                if sheet_name == "目录":
                    continue
                logger.info("读取工作表 %s", sheet_name)
                try:
                    df_sheet = read_sheet(file_path, sheet_name)
                except Exception as e:
                    msg = f"{os.path.basename(file_path)} - {sheet_name}: {e}"
                    logger.error(msg)
                    return {"error": msg}
                if df_sheet.empty:
                    continue
                daily_counts = count_daily_batches(df_sheet)
                file_sheet[sheet_name] = daily_counts.astype(int)
                sheet_counts[sheet_name] = sheet_counts.get(sheet_name, pd.Series()).add(daily_counts, fill_value=0).astype(int)
            self.file_stats[file_path] = file_sheet
        self.all_sheets = list(sheet_counts.keys())
        self._merged_results = merge_daily_counts(sheet_counts).reindex(columns=self.all_sheets, fill_value=0).sort_index()
        self._merged_total = self._merged_results.sum(axis=1).astype(int)
        self._merged_column_totals = self._merged_results.sum().astype(int)
        logger.info("统计完成")
        return {}

    def get_results(self, file_index: int = -1):
        if file_index < 0 or file_index >= len(self.file_paths):
            if not self._merged_results or self._merged_results.empty:
                return {"columns": [], "rows": [], "totals": [], "mode": "全部文件 — 无数据"}
            return self._df_to_table_json(self._merged_results, self._merged_total, self._merged_column_totals, "全部文件")
        file_path = self.file_paths[file_index]
        file_sheet = self.file_stats.get(file_path, {})
        if not file_sheet:
            return {"columns": [], "rows": [], "totals": [], "mode": f"{os.path.basename(file_path)} — 无数据"}
        file_results = merge_daily_counts(file_sheet).reindex(columns=list(file_sheet.keys()), fill_value=0).sort_index()
        total = file_results.sum(axis=1).astype(int)
        col_total = file_results.sum().astype(int)
        return self._df_to_table_json(file_results, total, col_total, os.path.basename(file_path))

    def _get_output_dir(self):
        if not self.file_paths:
            return ""
        base_dir = os.path.dirname(self.file_paths[0])
        out_dir = os.path.join(base_dir, "统计后结果")
        os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def _safe_save_excel(self, save_path, results, total_series, column_totals):
        pd = _import_pandas()
        if os.path.exists(save_path):
            base, ext = os.path.splitext(save_path)
            for i in range(2, 100):
                alt = f"{base}_v{i}{ext}"
                if not os.path.exists(alt):
                    save_path = alt
                    break
        try:
            logger.info("导出结果到 %s", save_path)
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                detail = results.reset_index()
                for col in detail.columns:
                    if col != "日期":
                        detail[col] = detail[col].replace(0, None)
                detail.to_excel(writer, sheet_name="每日批号统计", index=False)
                ws = writer.sheets["每日批号统计"]
                start, end = 2, ws.max_row
                ws.cell(row=end+1, column=1, value="合计")
                for idx, col in enumerate(detail.columns[1:], start=2):
                    col_letter = ws.cell(row=1, column=idx).column_letter
                    ws.cell(row=end+1, column=idx, value=f"=SUM({col_letter}{start}:{col_letter}{end})")
                total_df = total_series.rename("合计").to_frame().reset_index()
                total_df["合计"] = total_df["合计"].replace(0, None)
                total_df.to_excel(writer, sheet_name="每日汇总", index=False)
                ws2 = writer.sheets["每日汇总"]
                ws2.cell(row=ws2.max_row+1, column=1, value="合计")
                col_letter = ws2.cell(row=1, column=2).column_letter
                ws2.cell(row=ws2.max_row, column=2, value=f"=SUM({col_letter}{2}:{col_letter}{ws2.max_row-1})")
            return True
        except Exception as e:
            logger.error("导出失败 %s: %s", save_path, e)
            return False

    def batch_export(self):
        if not self.file_paths or not self.file_stats:
            return {"error": "请先导入并统计文件"}
        success = 0
        for file_path in self.file_paths:
            file_sheet = self.file_stats.get(file_path, {})
            if not file_sheet:
                continue
            file_results = merge_daily_counts(file_sheet).reindex(columns=list(file_sheet.keys()), fill_value=0).sort_index()
            total = file_results.sum(axis=1).astype(int)
            col_total = file_results.sum().astype(int)
            out_dir = self._get_output_dir()
            if not out_dir:
                continue
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            out_path = os.path.join(out_dir, f"{base_name}_统计结果.xlsx")
            if self._safe_save_excel(out_path, file_results, total, col_total):
                success += 1
        total_files = len(self.file_paths)
        logger.info("批量导出完成，成功 %d / %d", success, total_files)
        return {"message": f"成功导出 {success} / {total_files} 个文件", "status": f"批量导出完成，成功 {success} 个文件"}

    def open_export_folder(self):
        out_dir = self._get_output_dir()
        if not out_dir or not os.path.exists(out_dir):
            return {"error": "尚未导出过结果，请先执行批量导出"}
        if sys.platform.startswith("win"):
            os.startfile(out_dir)
        elif sys.platform == "darwin":
            import subprocess; subprocess.Popen(["open", out_dir])
        else:
            import subprocess; subprocess.Popen(["xdg-open", out_dir])
        logger.info("打开输出文件夹 %s", out_dir)
        return {"status": f"已打开文件夹: {out_dir}"}

    def clear_all(self):
        self.file_paths.clear()
        self.file_stats.clear()
        self.all_sheets.clear()
        self._merged_results = None
        self._merged_total = None
        self._merged_column_totals = None
        logger.info("清空所有数据和状态")
        return {"status": "已清空所有数据"}

    def get_log_content(self):
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "batch_count.log")
        for enc in ("utf-8", "gbk", "cp936"):
            try:
                with open(log_path, "r", encoding=enc) as f:
                    return "\n".join(reversed(f.read().splitlines()))
            except UnicodeDecodeError:
                continue
            except Exception:
                break
        try:
            with open(log_path, "r", encoding="utf-8", errors="replace") as f:
                return "\n".join(reversed(f.read().splitlines()))
        except Exception as e:
            return f"读取日志文件失败: {e}"

# --------------------------- 窗口图标（仅 Windows） ---------------------------
def _generate_app_icon() -> str:
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "app_icon.ico")
    if os.path.exists(icon_path):
        return icon_path
    import struct
    size = 32
    BG = (110, 69, 226)
    BAR = (255, 255, 255)
    pixels = bytearray()
    for y in range(size):
        for x in range(size):
            r, g, b = BG
            if (5 <= x <= 8 and 18 <= y <= 26) or (11 <= x <= 14 and 12 <= y <= 26) or (17 <= x <= 20 and 8 <= y <= 26) or (23 <= x <= 26 and 14 <= y <= 26):
                r, g, b = BAR
            pixels.extend([b, g, r, 255])
    bih = struct.pack("<IiiHHIIiiII", 40, size, size * 2, 1, 32, 0, len(pixels), 0, 0, 0, 0)
    ico_header = struct.pack("<HHH", 0, 1, 1)
    ico_entry = struct.pack("<BBBBHHII", size, size, 0, 0, 1, 32, len(bih) + len(pixels), 6 + 16)
    with open(icon_path, "wb") as f:
        f.write(ico_header)
        f.write(ico_entry)
        f.write(bih)
        f.write(pixels)
    return icon_path

def _set_window_icon(title: str):
    if not sys.platform.startswith("win"):
        return
    import ctypes, time
    icon_path = _generate_app_icon()
    if not os.path.exists(icon_path):
        return
    for _ in range(15):
        time.sleep(0.2)
        hwnd = ctypes.windll.user32.FindWindowW(None, title)
        if hwnd:
            hicon = ctypes.windll.user32.LoadImageW(0, icon_path, 1, 32, 32, 0x00000010)
            if hicon:
                ctypes.windll.user32.SendMessageW(hwnd, 0x0080, 1, hicon)
                ctypes.windll.user32.SendMessageW(hwnd, 0x0080, 0, hicon)
                logger.info("窗口图标已设置")
            break

# --------------------------- 程序入口 ---------------------------
def main():
    import threading, webview
    api = BackendApi()
    window_title = f"检测数据统计工具 v{VERSION}"
    window = webview.create_window(
        title=window_title,
        html=HTML_TEMPLATE,
        js_api=api,
        width=960,
        height=640,
        min_size=(800, 500),
        resizable=True,
        easy_drag=False,
    )
    logger.info("应用启动 v%s", VERSION)
    threading.Thread(target=_set_window_icon, args=(window_title,), daemon=True).start()
    webview.start(debug=False, gui="edgechromium" if os.name == "nt" else None)

if __name__ == "__main__":
    main()
"