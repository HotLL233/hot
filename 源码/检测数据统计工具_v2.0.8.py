#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
检测数据统计工具 v2.0.8 - Web UI 版本
基于 pywebview，使用 Uiverse Galaxy 风格的纯色按钮
-----------------------------------
功能：
1. 使用 pywebview 提供 Web 原生桌面界面
2. 按钮采用 Uiverse Galaxy 风格的纯色按钮（符合用户偏好）
3. 支持一次性导入多个 .xlsx 文件
4. 自动遍历每个工作表，统计每天每个工作表的批号去重数量
5. 汇总所有工作表的每日批号总数
6. 使用 pandas 读取 Excel，openpyxl 写入结果
7. 支持中文显示，具备基本错误处理
8. 暗色主题，现代化 UI
"""

# Version: 2.0.8
# 更新说明：
# - 修复重复的日期前向填充（两次 ffill 去重）
# - 修复 HTML_TEMPLATE 占位注释问题，恢复完整 UI
# - 正确更新 VERSION 常量
# - 保留全部 v2.0.7 优化功能

from __future__ import annotations
import os
import sys
import datetime as _dt
import logging
import json
import locale
locale.setlocale(locale.LC_ALL, '')

# --------------------------- 依赖检查 ---------------------------
try:
    import pandas as pd
except ModuleNotFoundError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd

try:
    import webview
except ModuleNotFoundError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pywebview"])
    import webview

from openpyxl.styles import Font, Alignment, Border, Side
from typing import Dict, Optional

# --------------------------- 日志配置 ---------------------------
import logging.handlers
def _resource_path(relative_path):
    """返回在 PyInstaller 打包环境或普通运行时的资源路径"""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

import os
log_dir = os.path.join(os.getenv('APPDATA') or os.path.expanduser('~'), '检测数据统计工具')
os.makedirs(log_dir, exist_ok=True)
log_file_path = os.path.join(log_dir, "batch_count.log")
_handler = logging.handlers.RotatingFileHandler(
    log_file_path, maxBytes=5 * 1024 * 1024, backupCount=3, delay=True, encoding="utf-8"
)
_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger(__name__)

# --------------------------- 配置区 ---------------------------
DATE_COL = "日期"
BATCH_COL = "批号"
SKIP_ROWS = 2
VERSION = "2.0.8"
AUTHOR = "HotLL"
RELEASE_DATE = "2026-05-21"

# --------------------------- 工具函数 (保留自 v1.5.2) ---------------------------

def _ensure_date_str(val) -> str:
    """统一转为 YYYY-MM-DD 字符串（处理 Excel 序列号、datetime、字符串）。"""
    if pd.isna(val):
        return ""
    if isinstance(val, (pd.Timestamp, _dt.datetime, _dt.date)):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("nat", "nan", "none", "null", "na"):
        return ""
    try:
        num = float(s)
        if 0 <= num <= 200000:
            dt = pd.Timestamp("1899-12-30") + pd.Timedelta(days=num)
            if 1900 <= dt.year <= 2100:
                return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    try:
        num = float(s)
        if 20250101 <= num <= 21000101:
            s_int = str(int(num))
            if len(s_int) == 8:
                dt = pd.to_datetime(s_int, format='%Y%m%d', errors='coerce')
                if pd.notna(dt):
                    return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    try:
        dt = pd.to_datetime(s, errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    return ""


def read_sheet(file_path: str, sheet_name: str) -> pd.DataFrame:
    """读取单个工作表，取前两列作为日期和批号。"""
    logger.info("读取工作表 %s 来自文件 %s", sheet_name, os.path.basename(file_path))
    try:
        df = pd.read_excel(
            file_path, sheet_name=sheet_name,
            header=None, skiprows=SKIP_ROWS,
            engine="openpyxl",
        )
    except Exception as e:
        logger.error("读取 %s - %s 失败: %s", os.path.basename(file_path), sheet_name, e)
        raise RuntimeError(f"读取 {os.path.basename(file_path)} - {sheet_name} 失败: {e}")
    if df.shape[1] < 2:
        return pd.DataFrame(columns=[DATE_COL, BATCH_COL])
    df = df.iloc[:, :2].copy()
    df.columns = [DATE_COL, BATCH_COL]
    df[DATE_COL] = df[DATE_COL].apply(_ensure_date_str)
    df[BATCH_COL] = df[BATCH_COL].astype(str).str.strip()
    # 前向填充日期一次即可（v2.0.8 优化：去掉重复 ffill）
    df[DATE_COL] = df[DATE_COL].replace("", pd.NA).ffill()
    df = df[df[DATE_COL].str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)]
    df = df[~df[BATCH_COL].str.startswith("例")]
    return df[[DATE_COL, BATCH_COL]]


def count_daily_batches(df: pd.DataFrame):
    """返回 Series：索引为日期，值为该日期批号去重计数。"""
    daily_counts = df.groupby(DATE_COL)[BATCH_COL].nunique()
    daily_counts.name = "count"
    return daily_counts


def merge_daily_counts(count_dict: Dict[str, pd.Series]) -> pd.DataFrame:
    """将各工作表的每日计数合并为宽表，列名为工作表名称。"""
    frames = []
    for sheet, series in count_dict.items():
        s = series.copy()
        s.name = sheet
        frames.append(s)
    if not frames:
        return pd.DataFrame()
    df_merged = pd.concat(frames, axis=1)
    df_merged = df_merged.fillna(0).astype(int)
    df_merged.index.name = DATE_COL
    return df_merged


# ========================== 前端 HTML ==========================
# 采用 Uiverse Galaxy 风格：暗色主题 + 纯色按钮 + 毛玻璃侧边栏

HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 32 32'><rect width='32' height='32' rx='6' fill='%236e45e2'/><rect x='5' y='18' width='4' height='9' rx='1' fill='white'/><rect x='11' y='12' width='4' height='15' rx='1' fill='white'/><rect x='17' y='8' width='4' height='19' rx='1' fill='white'/><rect x='23' y='14' width='4' height='13' rx='1' fill='white'/></svg>">
<title>检测数据统计工具 v2.0.8</title>
<style>
  /* ===================== 全局重置 ===================== */
  *, *::before, *::after {
    margin: 0;
    padding: 0;
    box-sizing: border-box;
  }

  :root {
    --bg-primary: #0d1117;
    --bg-secondary: #161b22;
    --bg-card: #1c2128;
    --bg-hover: #21262d;
    --border: #30363d;
    --text-primary: #e6edf3;
    --text-secondary: #8b949e;
    --text-dim: #6e7681;
    --purple-start: #6e45e2;
    --purple-end: #88d3ce;
    --blue-start: #2b5876;
    --blue-end: #4e4376;
    --green-start: #11998e;
    --green-end: #38ef7d;
    --red-start: #eb3349;
    --red-end: #f45c43;
    --gold-start: #f2994a;
    --gold-end: #f2c94c;
    --shadow-sm: 0 1px 3px rgba(0,0,0,0.3);
    --shadow-md: 0 4px 15px rgba(0,0,0,0.35);
    --shadow-lg: 0 8px 30px rgba(0,0,0,0.4);
    --radius-sm: 8px;
    --radius-md: 12px;
    --radius-lg: 16px;
    --transition: 0.3s cubic-bezier(0.23, 1, 0.32, 1);
  }

  body {
    font-family: 'Segoe UI', 'Microsoft YaHei UI', system-ui, sans-serif;
    background: var(--bg-primary);
    color: var(--text-primary);
    display: flex;
    height: 100vh;
    overflow: hidden;
    user-select: none;
    -webkit-user-select: none;
  }

  /* ===================== 侧边栏 ===================== */
  .sidebar {
    width: 200px;
    min-width: 200px;
    background: var(--bg-secondary);
    border-right: 1px solid var(--border);
    display: flex;
    flex-direction: column;
    padding: 24px 16px;
    gap: 10px;
    backdrop-filter: blur(10px);
  }

  .sidebar-title {
    font-size: 13px;
    font-weight: 700;
    color: var(--text-secondary);
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 6px;
    padding-left: 4px;
  }

  .sidebar-brand {
    font-size: 15px;
    font-weight: 700;
    color: #88d3ce;
    padding: 8px 12px;
    margin-bottom: 12px;
    background: rgba(110,69,226,0.12);
    border-radius: var(--radius-md);
    border: 1px solid rgba(110,69,226,0.2);
    text-align: center;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 8px;
  }

  .sidebar-brand .logo-icon {
    font-size: 22px;
    flex-shrink: 0;
  }

  .sidebar-brand span {
    color: #88d3ce;
    font-weight: 700;
  }

  /* ===================== 按钮 ===================== */
  .galaxy-btn {
    position: relative;
    width: 100%;
    padding: 11px 16px;
    border: none;
    border-radius: var(--radius-md);
    font-family: inherit;
    font-size: 13px;
    font-weight: 600;
    color: #fff;
    cursor: pointer;
    overflow: hidden;
    transition: transform var(--transition), box-shadow var(--transition);
    box-shadow: var(--shadow-sm);
    letter-spacing: 0.02em;
    text-align: left;
    display: flex;
    align-items: center;
    gap: 8px;
  }

  .galaxy-btn .btn-icon {
    font-size: 16px;
    width: 20px;
    text-align: center;
    flex-shrink: 0;
  }

  /* 按钮颜色变体 */
  .galaxy-btn.primary {
    background: #6e45e2;
  }
  .galaxy-btn.success {
    background: #11998e;
  }
  .galaxy-btn.info {
    background: #2b5876;
  }
  .galaxy-btn.danger {
    background: #eb3349;
  }
  .galaxy-btn.warning {
    background: #f2994a;
  }

  .galaxy-btn:disabled {
    opacity: 0.45;
    cursor: not-allowed;
    transform: none !important;
    box-shadow: none !important;
  }

  /* ===================== 主区域 ===================== */
  .main {
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow: hidden;
    padding: 20px;
    gap: 16px;
  }

  /* 顶部标题栏 */
  .header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding-bottom: 12px;
    border-bottom: 1px solid var(--border);
  }

  .header h1 {
    font-size: 20px;
    font-weight: 700;
    color: var(--text-primary);
  }

  .header-actions {
    display: flex;
    gap: 8px;
  }

  .header-btn {
    padding: 7px 14px;
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    background: var(--bg-card);
    color: var(--text-secondary);
    cursor: pointer;
    font-size: 12px;
    font-family: inherit;
    transition: all var(--transition);
  }

  .header-btn:hover {
    background: var(--bg-hover);
    color: var(--text-primary);
    border-color: var(--text-dim);
  }

  /* 文件列表面板 */
  .file-panel {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 16px;
    max-height: 180px;
    overflow: hidden;
    display: flex;
    flex-direction: column;
  }

  .panel-title {
    font-size: 13px;
    font-weight: 600;
    color: var(--text-secondary);
    margin-bottom: 10px;
    display: flex;
    align-items: center;
    gap: 6px;
  }

  .panel-title .count {
    background: var(--border);
    border-radius: 10px;
    padding: 1px 8px;
    font-size: 11px;
    color: var(--text-dim);
  }

  .file-list {
    list-style: none;
    overflow-y: auto;
    flex: 1;
    display: flex;
    flex-direction: column;
    gap: 3px;
  }

  .file-list::-webkit-scrollbar {
    width: 4px;
  }
  .file-list::-webkit-scrollbar-thumb {
    background: var(--border);
    border-radius: 2px;
  }

  .file-item {
    padding: 8px 12px;
    border-radius: var(--radius-sm);
    cursor: pointer;
    transition: all var(--transition);
    font-size: 13px;
    color: var(--text-secondary);
    display: flex;
    align-items: center;
    gap: 8px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }

  .file-item:hover {
    background: var(--bg-hover);
    color: var(--text-primary);
  }

  .file-item.active {
    background: rgba(110,69,226,0.15);
    color: #88d3ce;
    font-weight: 600;
    border: 1px solid rgba(110,69,226,0.25);
  }

  .file-item .file-icon {
    font-size: 15px;
    flex-shrink: 0;
  }

  .empty-state {
    color: var(--text-dim);
    font-size: 13px;
    text-align: center;
    padding: 20px;
  }

  /* 结果面板 */
  .result-panel {
    flex: 1;
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    overflow: hidden;
    display: flex;
    flex-direction: column;
  }

  .result-header {
    padding: 14px 16px;
    border-bottom: 1px solid var(--border);
    font-size: 13px;
    font-weight: 600;
    color: var(--text-secondary);
    display: flex;
    align-items: center;
    gap: 6px;
  }

  .table-wrapper {
    flex: 1;
    overflow: auto;
  }

  .table-wrapper::-webkit-scrollbar {
    width: 6px;
    height: 6px;
  }
  .table-wrapper::-webkit-scrollbar-thumb {
    background: var(--border);
    border-radius: 3px;
  }
  .table-wrapper::-webkit-scrollbar-track {
    background: transparent;
  }

  table {
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
  }

  thead {
    position: sticky;
    top: 0;
    z-index: 2;
  }

  thead th {
    background: var(--bg-secondary);
    padding: 10px 14px;
    text-align: center;
    font-weight: 600;
    color: var(--text-secondary);
    border-bottom: 2px solid var(--border);
    font-size: 12px;
    letter-spacing: 0.03em;
    white-space: nowrap;
  }

  tbody td {
    padding: 8px 14px;
    text-align: center;
    border-bottom: 1px solid rgba(48,54,61,0.4);
    color: var(--text-primary);
  }

  tbody tr:hover {
    background: rgba(110,69,226,0.06);
  }

  tbody tr:nth-child(even) {
    background: rgba(255,255,255,0.015);
  }

  tbody tr:nth-child(even):hover {
    background: rgba(110,69,226,0.08);
  }

  .row-total {
    font-weight: 700;
    background: rgba(110,69,226,0.08) !important;
    border-top: 2px solid rgba(110,69,226,0.3) !important;
  }

  .row-total td {
    font-weight: 700;
    color: #88d3ce;
  }

  .col-date {
    font-weight: 600;
    color: #a0a8c0;
    min-width: 100px;
  }

  /* 状态栏 */
  .status-bar {
    padding: 8px 16px;
    background: var(--bg-secondary);
    border-top: 1px solid var(--border);
    font-size: 12px;
    color: var(--text-dim);
    display: flex;
    align-items: center;
    gap: 8px;
    border-radius: 0 0 var(--radius-lg) var(--radius-lg);
  }

  .status-dot {
    width: 7px;
    height: 7px;
    border-radius: 50%;
    background: #38ef7d;
    flex-shrink: 0;
    animation: pulse 2s infinite;
  }

  .status-dot.busy {
    background: #f2c94c;
    animation: pulse 0.8s infinite;
  }

  .status-dot.error {
    background: #f45c43;
  }

  @keyframes pulse {
    0%, 100% { opacity: 1; }
    50% { opacity: 0.4; }
  }

  /* ===================== 弹窗 Modal ===================== */
  .modal-overlay {
    display: none;
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    background: rgba(0,0,0,0.6);
    z-index: 100;
    justify-content: center;
    align-items: center;
    backdrop-filter: blur(4px);
    -webkit-backdrop-filter: blur(4px);
  }

  .modal-overlay.active {
    display: flex;
  }

  .modal-box {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 24px;
    max-width: 700px;
    width: 90%;
    max-height: 80vh;
    display: flex;
    flex-direction: column;
    box-shadow: var(--shadow-lg);
    animation: modalIn 0.25s ease;
  }

  @keyframes modalIn {
    from { transform: scale(0.95) translateY(10px); opacity: 0; }
    to { transform: scale(1) translateY(0); opacity: 1; }
  }

  .modal-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 16px;
    padding-bottom: 12px;
    border-bottom: 1px solid var(--border);
  }

  .modal-header h3 {
    font-size: 16px;
    font-weight: 700;
  }

  .modal-close {
    background: none;
    border: none;
    color: var(--text-dim);
    font-size: 20px;
    cursor: pointer;
    padding: 4px 8px;
    border-radius: var(--radius-sm);
    transition: all var(--transition);
    line-height: 1;
  }

  .modal-close:hover {
    color: var(--text-primary);
    background: var(--bg-hover);
  }

  .modal-body {
    overflow-y: auto;
    flex: 1;
    font-size: 12px;
    font-family: 'Consolas', 'Courier New', 'Microsoft YaHei UI', monospace;
    color: var(--text-secondary);
    line-height: 1.6;
    white-space: pre-wrap;
  }

  .modal-body::-webkit-scrollbar {
    width: 4px;
  }
  .modal-body::-webkit-scrollbar-thumb {
    background: var(--border);
    border-radius: 2px;
  }

  .about-info {
    font-size: 14px;
    line-height: 2;
    color: var(--text-primary);
  }

  .about-info b {
    color: #88d3ce;
  }

  /* ===================== Toast 通知 ===================== */
  .toast-container {
    position: fixed;
    top: 20px;
    right: 20px;
    z-index: 200;
    display: flex;
    flex-direction: column;
    gap: 8px;
  }

  .toast {
    padding: 12px 20px;
    border-radius: var(--radius-md);
    font-size: 13px;
    font-weight: 500;
    color: #fff;
    box-shadow: var(--shadow-md);
    animation: toastIn 0.3s ease;
    max-width: 360px;
    word-break: break-word;
  }

  .toast.success {
    background: #11998e;
  }
  .toast.error {
    background: #eb3349;
  }
  .toast.info {
    background: #2b5876;
  }

  @keyframes toastIn {
    from { transform: translateX(100px); opacity: 0; }
    to { transform: translateX(0); opacity: 1; }
  }

  /* ===================== 浅色主题 ===================== */
  body.light {
    --bg-primary: #f6f8fa;
    --bg-secondary: #ffffff;
    --bg-card: #ffffff;
    --bg-hover: #f0f2f5;
    --border: #d0d7de;
    --text-primary: #1f2328;
    --text-secondary: #656d76;
    --text-dim: #8b949e;
    --shadow-sm: 0 1px 3px rgba(0,0,0,0.08);
    --shadow-md: 0 4px 15px rgba(0,0,0,0.1);
    --shadow-lg: 0 8px 30px rgba(0,0,0,0.12);
  }

  body.light thead th {
    background: #ffffff;
  }

  body.light tbody tr:nth-child(even) {
    background: rgba(0,0,0,0.02);
  }

  body.light tbody tr:nth-child(even):hover {
    background: rgba(110,69,226,0.04);
  }

  body.light .row-total {
    background: rgba(110,69,226,0.04) !important;
  }

  body.light .file-item.active {
    background: rgba(110,69,226,0.06);
    color: #6e45e2;
    border: 1px solid rgba(110,69,226,0.2);
  }

  body.light .header-btn {
    background: #f6f8fa;
  }

  body.light .header-btn:hover {
    background: #e8eaed;
  }

  body.light .sidebar-brand {
    background: rgba(110,69,226,0.06);
    border: 1px solid rgba(110,69,226,0.12);
    color: #6e45e2;
  }

  body.light .sidebar-brand span {
    color: #6e45e2;
  }

  .theme-toggle {
    background: none;
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    padding: 6px 10px;
    cursor: pointer;
    font-size: 16px;
    transition: all var(--transition);
    color: var(--text-secondary);
    line-height: 1;
  }

  .theme-toggle:hover {
    background: var(--bg-hover);
    color: var(--text-primary);
  }

  /* ===================== 确认弹窗 ===================== */
  .confirm-modal .modal-box {
    max-width: 420px;
    text-align: center;
  }

  .confirm-modal .confirm-icon {
    font-size: 48px;
    margin-bottom: 8px;
    line-height: 1;
  }

  .confirm-modal .confirm-msg {
    font-size: 14px;
    color: var(--text-primary);
    margin-bottom: 20px;
    line-height: 1.6;
  }

  .confirm-actions {
    display: flex;
    gap: 12px;
    justify-content: center;
  }

  .confirm-btn {
    padding: 9px 28px;
    border: none;
    border-radius: var(--radius-md);
    font-family: inherit;
    font-size: 13px;
    font-weight: 600;
    cursor: pointer;
    transition: all var(--transition);
    min-width: 90px;
  }

  .confirm-btn.ok {
    background: #11998e;
    color: #fff;
  }

  .confirm-btn.ok:hover {
    box-shadow: 0 2px 10px rgba(17, 153, 142, 0.35);
    transform: translateY(-1px);
  }

  .confirm-btn.cancel {
    background: var(--bg-hover);
    color: var(--text-secondary);
    border: 1px solid var(--border);
  }

  .confirm-btn.cancel:hover {
    background: var(--border);
    color: var(--text-primary);
  }

  /* ===================== 响应式 ===================== */
  @media (max-width: 600px) {
    .sidebar {
      width: 56px;
      min-width: 56px;
      padding: 16px 6px;
    }
    .galaxy-btn {
      font-size: 0;
      justify-content: center;
      padding: 10px;
    }
    .galaxy-btn .btn-icon {
      margin: 0;
    }
    .sidebar-title, .sidebar-brand {
      display: none;
    }
  }
</style>
</head>
<body class="light">

  <!-- ========== 侧边栏 ========== -->
  <aside class="sidebar">
    <div class="sidebar-brand">
      <span class="logo-icon">&#128202;</span><span>检测数据统计</span>
    </div>
    <div class="sidebar-title">操作</div>
    <button class="galaxy-btn primary" onclick="importFiles()">
      <span class="btn-icon">&#128194;</span> 导入文件
    </button>
    <button class="galaxy-btn success" onclick="batchExport()">
      <span class="btn-icon">&#128190;</span> 批量导出
    </button>
    <button class="galaxy-btn info" onclick="openExportFolder()">
      <span class="btn-icon">&#128193;</span> 打开输出目录
    </button>
    <button class="galaxy-btn danger" onclick="clearAll()">
      <span class="btn-icon">&#128465;</span> 清空列表
    </button>
  </aside>

  <!-- ========== 主区域 ========== -->
  <div class="main">
    <!-- 标题栏 -->
    <div class="header">
      <h1>检测数据统计工具</h1>
      <div class="header-actions">
        <button class="theme-toggle" onclick="toggleTheme()" title="切换主题" id="theme-btn">&#9788;</button>
        <button class="header-btn" onclick="showAbout()">关于</button>
        <button class="header-btn" onclick="showLog()">审计日志</button>
      </div>
    </div>

    <!-- 文件列表 -->
    <div class="file-panel">
      <div class="panel-title">
        &#128196; 已导入文件 <span class="count" id="file-count">0</span>
      </div>
      <ul class="file-list" id="file-list">
        <li class="empty-state">暂无文件，请点击左侧「导入文件」</li>
      </ul>
    </div>

    <!-- 统计结果 -->
    <div class="result-panel">
      <div class="result-header">
        &#128202; 统计结果预览 <span id="view-mode" style="font-weight:400;color:var(--text-dim);font-size:11px;">— 全部文件</span>
      </div>
      <div class="table-wrapper">
        <table id="result-table">
          <thead></thead>
          <tbody></tbody>
        </table>
      </div>
      <div class="status-bar">
        <span class="status-dot" id="status-dot"></span>
        <span id="status-text">就绪</span>
      </div>
    </div>
  </div>

  <!-- ========== 确认弹窗 ========== -->
  <div class="modal-overlay confirm-modal" id="confirm-modal">
    <div class="modal-box">
      <div class="confirm-icon">&#128204;</div>
      <div class="confirm-msg" id="confirm-msg">确认操作？</div>
      <div class="confirm-actions">
        <button class="confirm-btn ok" id="confirm-ok" onclick="confirmAction(true)">确定</button>
        <button class="confirm-btn cancel" onclick="confirmAction(false)">取消</button>
      </div>
    </div>
  </div>

  <!-- ========== Toast 容器 ========== -->
  <div class="toast-container" id="toast-container"></div>

  <!-- ========== 审计日志弹窗 ========== -->
  <div class="modal-overlay" id="log-modal">
    <div class="modal-box">
      <div class="modal-header">
        <h3>&#128220; 审计日志</h3>
        <div style="display:flex;gap:8px;">
          <button class="header-btn" onclick="refreshLog()">刷新</button>
          <button class="modal-close" onclick="closeModal('log-modal')">&times;</button>
        </div>
      </div>
      <div class="modal-body" id="log-content"></div>
    </div>
  </div>

  <!-- ========== 关于弹窗 ========== -->
  <div class="modal-overlay" id="about-modal">
    <div class="modal-box" style="max-width:400px;">
      <div class="modal-header">
        <h3>&#8505; 关于</h3>
        <button class="modal-close" onclick="closeModal('about-modal')">&times;</button>
      </div>
      <div class="about-info">
        版本: <b>__VERSION__</b><br>
        作者: <b>__AUTHOR__</b><br>
        发布日期: <b>__RELEASE_DATE__</b><br><br>
        基于 pywebview + Uiverse Galaxy UI
      </div>
    </div>
  </div>

  <script>
    // ===================== 全局状态 =====================
    let selectedFileIndex = -1;  // -1 = 全部文件
    let fileList = [];

    // ===================== 工具函数 =====================
    function setStatus(msg, type) {{
      const text = document.getElementById('status-text');
      const dot = document.getElementById('status-dot');
      text.textContent = msg;
      dot.className = 'status-dot';
      if (type === 'busy') dot.classList.add('busy');
      if (type === 'error') dot.classList.add('error');
    }}

    function showToast(msg, type) {{
      const container = document.getElementById('toast-container');
      const toast = document.createElement('div');
      toast.className = 'toast ' + (type || 'info');
      toast.textContent = msg;
      container.appendChild(toast);
      setTimeout(() => toast.remove(), 3500);
    }}

    function closeModal(id) {{
      document.getElementById(id).classList.remove('active');
    }}

    function openModal(id) {{
      document.getElementById(id).classList.add('active');
    }}

    // ===================== 文件列表 =====================
    function updateFileList(files) {{
      fileList = files;
      const ul = document.getElementById('file-list');
      const count = document.getElementById('file-count');
      count.textContent = files.length;

      if (files.length === 0) {{
        ul.innerHTML = '<li class="empty-state">暂无文件，请点击左侧「导入文件」</li>';
        selectedFileIndex = -1;
        return;
      }}

      let html = '';
      files.forEach((name, i) => {{
        const cls = (i === selectedFileIndex) ? 'file-item active' : 'file-item';
        html += `<li class="${{cls}}" onclick="selectFile(${{i}})" title="${{name}}">
          <span class="file-icon">&#128203;</span> ${{name}}
        </li>`;
      }});
      ul.innerHTML = html;

      // 滚动到选中项
      if (selectedFileIndex >= 0) {{
        const items = ul.querySelectorAll('.file-item');
        if (items[selectedFileIndex]) items[selectedFileIndex].scrollIntoView({{ block: 'nearest' }});
      }}
    }}

    // ===================== 表格渲染 =====================
    function renderTable(data) {{
      const thead = document.querySelector('#result-table thead');
      const tbody = document.querySelector('#result-table tbody');
      const modeLabel = document.getElementById('view-mode');

      if (!data || !data.columns || data.columns.length === 0) {{
        thead.innerHTML = '';
        tbody.innerHTML = '<tr><td colspan="10" style="padding:40px;color:var(--text-dim);">暂无统计数据</td></tr>';
        modeLabel.textContent = '— 无数据';
        return;
      }}

      // 表头
      thead.innerHTML = '<tr>' + data.columns.map(c => `<th>${{c}}</th>`).join('') + '</tr>';

      // 数据行
      let bodyHtml = '';
      const rows = data.rows || [];
      rows.forEach((row, ri) => {{
        bodyHtml += '<tr>';
        row.forEach((cell, ci) => {{
          const cn = ci === 0 ? 'col-date' : '';
          bodyHtml += `<td class="${{cn}}">${{cell !== null && cell !== undefined ? cell : ''}}</td>`;
        }});
        bodyHtml += '</tr>';
      }});

      // 合计行
      if (data.totals && data.totals.length > 0) {{
        bodyHtml += '<tr class="row-total">';
        data.totals.forEach((cell, ci) => {{
          const cn = ci === 0 ? 'col-date' : '';
          bodyHtml += `<td class="${{cn}}">${{cell !== null && cell !== undefined ? cell : ''}}</td>`;
        }});
        bodyHtml += '</tr>';
      }}

      tbody.innerHTML = bodyHtml;

      // 视图模式
      if (data.mode) {{
        modeLabel.textContent = '— ' + data.mode;
      }}
    }}

    // ===================== API 调用 =====================
    async function importFiles() {{
      setStatus('正在打开文件选择…', 'busy');
      try {{
        const result = await pywebview.api.import_files();
        if (result.error) {{
          showToast(result.error, 'error');
          setStatus('导入失败', 'error');
          return;
        }}
        updateFileList(result.files || []);
        if (result.results) {{
          renderTable(result.results);
          selectedFileIndex = -1;
        }}
        setStatus(result.status || '就绪');
        if (result.files && result.files.length > 0) {{
          showToast('成功导入 ' + result.files.length + ' 个文件', 'success');
        }}
      }} catch (e) {{
        showToast('导入出错: ' + e, 'error');
        setStatus('导入出错', 'error');
      }}
    }}

    async function selectFile(index) {{
      if (index === selectedFileIndex) {{
        // 再次点击同文件 -> 回到全部
        selectedFileIndex = -1;
      }} else {{
        selectedFileIndex = index;
      }}
      updateFileList(fileList);
      setStatus('加载中…', 'busy');
      try {{
        const data = await pywebview.api.get_results(selectedFileIndex);
        renderTable(data);
        setStatus(data.mode || '就绪');
      }} catch (e) {{
        showToast('加载失败: ' + e, 'error');
        setStatus('加载失败', 'error');
      }}
    }}

    function batchExport() {{
      if (fileList.length === 0) {{
        showToast('请先导入文件', 'info');
        return;
      }}
      document.getElementById('confirm-msg').textContent =
        `将为 ${fileList.length} 个文件分别生成统计结果，是否继续？`;
      openModal('confirm-modal');
    }}

    function confirmAction(confirmed) {{
      closeModal('confirm-modal');
      if (!confirmed) return;
      setStatus('批量导出中…', 'busy');
      (async () => {{
        try {{
          const result = await pywebview.api.batch_export();
          if (result.error) {{
            showToast(result.error, 'error');
            setStatus('导出失败', 'error');
            return;
          }}
          showToast(result.message || '导出完成', 'success');
          setStatus(result.status || '就绪');
        }} catch (e) {{
          showToast('导出出错: ' + e, 'error');
          setStatus('导出出错', 'error');
        }}
      }})();
    }}

    async function openExportFolder() {{
      try {{
        const result = await pywebview.api.open_export_folder();
        if (result.error) {{
          showToast(result.error, 'info');
        }} else {{
          setStatus(result.status || '已打开文件夹');
        }}
      }} catch (e) {{
        showToast('操作失败: ' + e, 'error');
      }}
    }}

    async function clearAll() {{
      try {{
        const result = await pywebview.api.clear_all();
        updateFileList([]);
        renderTable(null);
        selectedFileIndex = -1;
        setStatus(result.status || '已清空');
      }} catch (e) {{
        showToast('清空失败: ' + e, 'error');
      }}
    }}

    async function showLog() {{
      openModal('log-modal');
      setStatus('加载日志中…', 'busy');
      try {{
        const content = await pywebview.api.get_log_content();
        document.getElementById('log-content').textContent = content;
        setStatus('就绪');
      }} catch (e) {{
        document.getElementById('log-content').textContent = '读取日志失败: ' + e;
        setStatus('读取日志失败', 'error');
      }}
    }}

    async function refreshLog() {{
      try {{
        const content = await pywebview.api.get_log_content();
        document.getElementById('log-content').textContent = content;
      }} catch (e) {{
        document.getElementById('log-content').textContent = '读取日志失败: ' + e;
      }}
    }}

    function showAbout() {{
      openModal('about-modal');
    }}

    // ===================== 点击遮罩关闭弹窗 =====================
    document.querySelectorAll('.modal-overlay').forEach(overlay => {{
      overlay.addEventListener('click', function(e) {{
        if (e.target === this) closeModal(this.id);
      }});
    }});

    // ===================== ESC 关闭弹窗 =====================
    document.addEventListener('keydown', function(e) {{
      if (e.key === 'Escape') {{
        document.querySelectorAll('.modal-overlay.active').forEach(m => m.classList.remove('active'));
      }}
    }});

    // ===================== 初始化 =====================
    setStatus('就绪');
    initTheme();

    // ===================== 主题切换 =====================
    function initTheme() {
      var saved = localStorage.getItem('app-theme');
      if (saved !== 'dark') {
        document.getElementById('theme-btn').innerHTML = '&#9788;';
      } else {
        document.body.classList.remove('light');
        document.getElementById('theme-btn').innerHTML = '&#9789;';
      }
    }

    function toggleTheme() {
      var body = document.body;
      var btn = document.getElementById('theme-btn');
      body.classList.toggle('light');
      if (body.classList.contains('light')) {
        localStorage.setItem('app-theme', 'light');
        btn.innerHTML = '&#9788;';
      } else {
        localStorage.setItem('app-theme', 'dark');
        btn.innerHTML = '&#9789;';
      }
    }
  </script>
</body>
</html>
"""
# 将 JS 中为 .format() 转义的双花括号还原为单花括号
HTML_TEMPLATE = HTML_TEMPLATE.replace("{{", "{").replace("}}", "}")
HTML_TEMPLATE = HTML_TEMPLATE.replace("__VERSION__", VERSION)
HTML_TEMPLATE = HTML_TEMPLATE.replace("__AUTHOR__", AUTHOR)
HTML_TEMPLATE = HTML_TEMPLATE.replace("__RELEASE_DATE__", RELEASE_DATE)


# ========================== Python 后端 API ==========================

class BackendApi:
    """暴露给前端 JS 的 API 类，通过 pywebview 的 js_api 桥接"""

    def __init__(self):
        self.file_paths: list[str] = []
        self.file_stats: dict = {}
        self.all_sheets: list[str] = []
        self._merged_results: Optional[pd.DataFrame] = None
        self._merged_total: Optional[pd.Series] = None
        self._merged_column_totals: Optional[pd.Series] = None

    # ---------- 结果转 JSON ----------
    def _df_to_table_json(self, results: pd.DataFrame, total_series: pd.Series,
                          column_totals: pd.Series, mode: str = "") -> dict:
        """将 DataFrame 转换为前端表格 JSON"""
        if results.empty:
            return {"columns": [], "rows": [], "totals": [], "mode": mode}
        columns = ["日期"] + list(results.columns) + ["合计"]
        num = results.fillna(0).astype(int)
        rows = [[date] + list(row[1:]) + [int(total_series.get(date, 0))]
                for date, row in zip(num.index, num.itertuples())]
        totals = ["合计"] + [int(column_totals.get(c, 0)) for c in results.columns] + [int(column_totals.sum())]
        return {"columns": columns, "rows": rows, "totals": totals, "mode": mode}

    # ---------- 导入文件 ----------
    def import_files(self) -> dict:
        """打开文件选择对话框，导入并处理 Excel 文件"""
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        paths = filedialog.askopenfilenames(
            title="请选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        root.destroy()

        if not paths:
            return {"error": "未选择任何文件", "files": [], "results": None, "status": "就绪"}

        self.file_paths = list(paths)
        logger.info("导入 %d 个文件: %s", len(self.file_paths),
                    [os.path.basename(p) for p in self.file_paths])

        # 处理文件
        result = self._process_files()
        if result.get("error"):
            return result

        file_names = [os.path.basename(p) for p in self.file_paths]
        return {
            "files": file_names,
            "results": self._df_to_table_json(
                self._merged_results, self._merged_total,
                self._merged_column_totals, "全部文件"
            ),
            "status": f"已导入 {len(self.file_paths)} 个文件，统计完成",
        }

    def _process_files(self) -> dict:
        """处理所有已导入文件（顺序执行 + 逐表容错）"""
        if not self.file_paths:
            return {"error": "请先导入文件"}

        logger.info("开始处理 %d 个文件", len(self.file_paths))
        sheet_counts = {}
        self.file_stats.clear()

        for file_path in self.file_paths:
            logger.info("处理文件 %s", os.path.basename(file_path))
            try:
                xl = pd.ExcelFile(file_path, engine="openpyxl")
            except Exception as e:
                logger.error("文件 %s 读取失败(跳过): %s", os.path.basename(file_path), e)
                continue

            file_sheet = {}
            for sheet_name in xl.sheet_names:
                if sheet_name == "目录":
                    continue
                try:
                    df_sheet = read_sheet(file_path, sheet_name)
                except Exception as e:
                    logger.error("%s - %s 读取失败(跳过): %s", os.path.basename(file_path), sheet_name, e)
                    continue
                if df_sheet.empty:
                    continue
                daily_counts = count_daily_batches(df_sheet).astype(int)
                file_sheet[sheet_name] = daily_counts
                if sheet_name in sheet_counts:
                    sheet_counts[sheet_name] = sheet_counts[sheet_name].add(daily_counts, fill_value=0).astype(int)
                else:
                    sheet_counts[sheet_name] = daily_counts
            if file_sheet:
                self.file_stats[file_path] = file_sheet

        self.all_sheets = list(sheet_counts.keys())
        results = merge_daily_counts(sheet_counts)
        results = results.reindex(columns=self.all_sheets, fill_value=0)
        results.sort_index(inplace=True)

        if results.empty:
            logger.warning("统计结果为空")
            return {"error": "未检测到有效数据"}

        self._merged_results = results.copy()
        self._merged_total = results.sum(axis=1).astype(int)
        self._merged_column_totals = results.sum().astype(int)
        logger.info("统计完成")
        return {}

    # ---------- 获取结果 ----------
    def get_results(self, file_index: int = -1) -> dict:
        """获取统计结果，-1 表示全部汇总，>=0 表示单个文件"""
        if file_index < 0 or file_index >= len(self.file_paths):
            # 全部汇总
            if self._merged_results is None or self._merged_results.empty:
                return {"columns": [], "rows": [], "totals": [], "mode": "全部文件 — 无数据"}
            return self._df_to_table_json(
                self._merged_results, self._merged_total,
                self._merged_column_totals, "全部文件"
            )

        file_path = self.file_paths[file_index]
        file_sheet = self.file_stats.get(file_path, {})
        if not file_sheet:
            return {"columns": [], "rows": [], "totals": [],
                    "mode": f"{os.path.basename(file_path)} — 无数据"}

        file_results = merge_daily_counts(file_sheet)
        file_sheet_names = list(file_sheet.keys())
        file_results = file_results.reindex(columns=file_sheet_names, fill_value=0)
        file_results.sort_index(inplace=True)
        file_total = file_results.sum(axis=1).astype(int)
        file_col_total = file_results.sum().astype(int)
        return self._df_to_table_json(
            file_results, file_total, file_col_total,
            f"{os.path.basename(file_path)}"
        )

    # ---------- 批量导出 ----------
    def batch_export(self) -> dict:
        """批量导出所有文件的统计结果"""
        if not self.file_paths or not self.file_stats:
            return {"error": "请先导入并统计文件"}

        # 使用 JS confirm（前端的 confirm 对话框）
        # 这里直接执行导出
        success_count = 0
        for file_path in self.file_paths:
            file_sheet = self.file_stats.get(file_path, {})
            if not file_sheet:
                continue
            file_results = merge_daily_counts(file_sheet)
            file_sheet_names = list(file_sheet.keys())
            file_results = file_results.reindex(columns=file_sheet_names, fill_value=0)
            file_results.sort_index(inplace=True)
            file_total = file_results.sum(axis=1).astype(int)
            file_col_total = file_results.sum().astype(int)

            out_dir = self._get_output_dir()
            if not out_dir:
                continue
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            out_path = os.path.join(out_dir, f"{base_name}_统计结果.xlsx")
            ok = self._write_formatted_excel(
                out_path, file_results, file_total, file_col_total
            )
            if ok:
                success_count += 1
                logger.info("批量导出成功: %s", out_path)

        total = len(self.file_paths)
        logger.info("批量导出完成，成功 %d / %d 个文件", success_count, total)
        return {
            "message": f"成功导出 {success_count} / {total} 个文件",
            "status": f"批量导出完成，成功 {success_count} 个文件",
        }

    def _get_output_dir(self) -> str:
        """返回持久化的导出目录（与日志同目录）"""
        if not self.file_paths:
            return ""
        # 使用与日志相同的 APPDATA 目录，确保跨机器、跨用户都有写权限
        log_dir = os.path.join(os.getenv('APPDATA') or os.path.expanduser('~'), '检测数据统计工具')
        out_dir = os.path.join(log_dir, '统计后结果')
        os.makedirs(out_dir, exist_ok=True)
        return out_dir

    def _write_formatted_excel(self, save_path, results, total_series, column_totals) -> bool:
        """写入格式化 Excel，文件占用时自动添加版本号后缀"""
        if os.path.exists(save_path):
            try:
                os.remove(save_path)
            except PermissionError:
                base, ext = os.path.splitext(save_path)
                for ver in range(2, 100):
                    alt_path = f"{base}_v{ver}{ext}"
                    if not os.path.exists(alt_path):
                        logger.warning("导出文件被占用 %s，自动改名为 %s", save_path, alt_path)
                        save_path = alt_path
                        break
                    try:
                        os.remove(alt_path)
                        save_path = alt_path
                        break
                    except PermissionError:
                        continue
                else:
                    logger.warning("导出文件被占用且无法改名 %s", save_path)
                    return False
        try:
            logger.info("导出结果到 %s", save_path)
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                detail_df = results.reset_index()
                for col in detail_df.columns:
                    if col != "日期":
                        detail_df[col] = detail_df[col].replace(0, None)
                detail_df.to_excel(writer, sheet_name="每日批号统计", index=False)
                ws_detail = writer.sheets["每日批号统计"]
                start_row = 2
                end_row = ws_detail.max_row
                total_row = end_row + 1
                ws_detail.cell(row=total_row, column=1, value="合计")
                for col_idx, col_name in enumerate(detail_df.columns[1:], start=2):
                    col_letter = ws_detail.cell(row=1, column=col_idx).column_letter
                    formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                    ws_detail.cell(row=total_row, column=col_idx, value=formula)

                total_df = total_series.rename("合计").to_frame().reset_index()
                total_df["合计"] = total_df["合计"].replace(0, None)
                total_df.to_excel(writer, sheet_name="每日汇总", index=False)
                ws_total = writer.sheets["每日汇总"]
                start_row = 2
                end_row = ws_total.max_row
                total_row = end_row + 1
                ws_total.cell(row=total_row, column=1, value="合计")
                col_letter = ws_total.cell(row=1, column=2).column_letter
                formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                ws_total.cell(row=total_row, column=2, value=formula)

                font_style = Font(name="仿宋", size=14)
                align_center = Alignment(horizontal="center", vertical="center")
                for ws in writer.sheets.values():
                    thin = Side(style="thin")
                    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    total_row = ws.max_row
                    total_col = ws.max_column
                    for col_cells in ws.columns:
                        max_len = 0
                        col_letter = col_cells[0].column_letter
                        for cell in col_cells:
                            cell.font = font_style
                            cell.alignment = align_center
                            cell.border = thin_border
                            if cell.row == total_row or cell.column == total_col:
                                cell.font = Font(name="仿宋", size=14, bold=True)
                            val = str(cell.value) if cell.value is not None else ""
                            char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                            max_len = max(max_len, char_len + 2)
                        ws.column_dimensions[col_letter].width = max(
                            min(max_len + 2, 30), 8
                        )
                    for row_dim in ws.row_dimensions.values():
                        row_dim.height = 35
            logger.info("导出成功 %s", save_path)
            return True
        except Exception as e:
            logger.error("导出失败 %s: %s", save_path, e)
            return False

    # ---------- 打开输出目录 ----------
    def open_export_folder(self) -> dict:
        out_dir = self._get_output_dir()
        if not out_dir or not os.path.exists(out_dir):
            return {"error": "尚未导出过结果，请先执行批量导出"}
        import subprocess, platform
        if platform.system() == "Windows":
            os.startfile(out_dir)
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", out_dir])
        else:
            subprocess.Popen(["xdg-open", out_dir])
        logger.info("打开输出文件夹 %s", out_dir)
        return {"status": f"已打开文件夹: {out_dir}"}

    # ---------- 清空 ----------
    def clear_all(self) -> dict:
        self.file_paths.clear()
        self.file_stats.clear()
        self.all_sheets.clear()
        self._merged_results = None
        self._merged_total = None
        self._merged_column_totals = None
        logger.info("清空所有数据和状态")
        return {"status": "已清空所有数据"}

    # ---------- 审计日志 ----------
    def get_log_content(self) -> str:
        """读取审计日志尾部（最多 500 行），降序返回"""
        import collections
        MAX_LINES = 500
        for enc in ("utf-8", "gbk", "cp936"):
            try:
                with open(log_file_path, "r", encoding=enc) as f:
                    lines = collections.deque(f, maxlen=MAX_LINES)
                    return "\n".join(reversed(lines))
            except UnicodeDecodeError:
                continue
            except Exception:
                break
        try:
            with open(log_file_path, "r", encoding="utf-8", errors="replace") as f:
                lines = collections.deque(f, maxlen=MAX_LINES)
                return "\n".join(reversed(lines))
        except Exception as e:
            return f"读取日志文件失败: {e}"


# ========================== 主入口 ==========================

def _generate_app_icon() -> str:
    """生成应用程序图标（柱状图 .ico 文件），返回图标路径"""
    import struct
    icon_path = _resource_path("app_icon.ico")
    if os.path.exists(icon_path):
        return icon_path

    # 画 32x32 柱状图：紫底 + 4 条白柱
    size = 32
    BG = (110, 69, 226)   # #6e45e2
    BAR = (255, 255, 255)
    pixels = bytearray()
    for y in range(size):
        for x in range(size):
            r, g, b = BG
            # 柱状图区域
            if (5 <= x <= 8 and 18 <= y <= 26) or \
               (11 <= x <= 14 and 12 <= y <= 26) or \
               (17 <= x <= 20 and 8 <= y <= 26) or \
               (23 <= x <= 26 and 14 <= y <= 26):
                r, g, b = BAR
            pixels.extend([b, g, r, 255])  # BGRA

    # BITMAPINFOHEADER
    bih = struct.pack("<IiiHHIIiiII",
        40, size, size * 2, 1, 32, 0,
        len(pixels), 0, 0, 0, 0)

    # ICO header
    num_images = 1
    data_offset = 6 + 16
    ico_header = struct.pack("<HHH", 0, 1, num_images)
    ico_entry = struct.pack("<BBBBHHII",
        size, size, 0, 0, 1, 32,
        len(bih) + len(pixels), data_offset)

    with open(icon_path, "wb") as f:
        f.write(ico_header)
        f.write(ico_entry)
        f.write(bih)
        f.write(bytes(pixels))
    return icon_path


def _set_window_icon(title: str):
    """通过 Windows API 设置窗口图标（单次延迟后尝试）"""
    if os.name != "nt":
        return
    import ctypes, time
    icon_path = _generate_app_icon()
    if not os.path.exists(icon_path):
        return
    time.sleep(1.0)  # 等待窗口就绪
    hwnd = ctypes.windll.user32.FindWindowW(None, title)
    if hwnd:
        icon_handle = ctypes.windll.user32.LoadImageW(0, icon_path, 1, 32, 32, 0x10)
        if icon_handle:
            ctypes.windll.user32.SendMessageW(hwnd, 0x0080, 1, icon_handle)
            ctypes.windll.user32.SendMessageW(hwnd, 0x0080, 0, icon_handle)
            logger.info("窗口图标已设置")


def main():
    """启动 pywebview 桌面应用"""
    import threading
    api = BackendApi()
    window_title = "检测数据统计工具 v2.0.8"
    window = webview.create_window(
        title=window_title,
        html=HTML_TEMPLATE,
        js_api=api,
        width=960,
        height=640,
        min_size=(800, 500),
        resizable=True,
        easy_drag=False,
    )
    logger.info("应用启动 v2.0.8")
    threading.Thread(target=_set_window_icon, args=(window_title,), daemon=True).start()
    webview.start(debug=False, gui="edgechromium" if os.name == "nt" else None)


if __name__ == "__main__":
    main()
